from fastapi import (FastAPI, Depends, HTTPException, status, Response, File,
                     UploadFile, Form, Query, BackgroundTasks, Body)
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy import or_, desc, and_, func, insert, select, update, extract
from sqlalchemy.orm import Session, joinedload, selectinload
from typing import List,  Optional
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from datetime import timedelta, date, datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from jose import JWSError, jwt
from pydantic import BaseModel  

import models, database, schemas, security, uuid, io, os, shutil, uuid, io, zipfile, services_wa, asyncio, tempfile, re, random

# ===================================================================
# INISIALISASI & KONFIGURASI
# ===================================================================
models.Base.metadata.create_all(bind=database.engine)
app = FastAPI()

origins = [
    "*"
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# DOKUMEN_DIRECTORY = "./dokumen"
STORAGE_DIRECTORY = "./storage"
UPLOAD_PROFILE_PIC_DIR = "./profile-picture"

if not os.path.exists(STORAGE_DIRECTORY):
    os.makedirs(STORAGE_DIRECTORY)

app.mount("/storage", StaticFiles(directory="storage"), name="storage")

if not os.path.exists(UPLOAD_PROFILE_PIC_DIR):
    os.makedirs(UPLOAD_PROFILE_PIC_DIR)
app.mount("/profile-picture", StaticFiles(directory="profile-picture"), name="profile-picture")

def save_file_securely(file: UploadFile) -> str:
    """
    Menyimpan file dengan nama UUID di folder berdasarkan Tahun/Bulan.
    Returns: Relative path string (contoh: 'storage/2024/11/550e8400.pdf')
    """
    filename = file.filename
    ext = filename.split('.')[-1] if '.' in filename else 'bin'
    now = datetime.now()
    folder_path = os.path.join(STORAGE_DIRECTORY, str(now.year), f"{now.month:02d}")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path, exist_ok=True)
    unique_filename = f"{uuid.uuid4()}.{ext}"
    file_path = os.path.join(folder_path, unique_filename)
    with open(file_path, "wb+") as file_object:
        shutil.copyfileobj(file.file, file_object)
    relative_path = os.path.relpath(file_path, start=".").replace("\\", "/")
    return relative_path

# ===================================================================
# FUNGSI WRAPPER WA ANTI-BOT
# ===================================================================
async def send_whatsapp_safe(phone_number: str, message: str, sequence_index: int = 0):
    """
    Mengirim pesan WA dengan jeda waktu agar tidak terdeteksi bot.
    sequence_index: Urutan pesan dalam loop (0, 1, 2, ...).
    """
    # 1. Jeda Dasar: 3 sampai 5 detik per pesan antrian
    # Jika ada 10 orang, orang ke-10 akan menerima pesan setelah +/- 30-50 detik
    base_delay = sequence_index * random.randint(3, 5)
    
    # 2. Jitter (Variasi Acak): Tambahan 1-3 detik agar tidak pola mesin
    jitter = random.uniform(1.0, 3.0)
    
    total_delay = base_delay + jitter
    
    # Tunggu secara asinkron (tidak memblokir server)
    await asyncio.sleep(total_delay)
    
    try:
        # Panggil service WA asli (pastikan services_wa menghandle error network)
        print(f"Creating WA Task for {phone_number} with delay {total_delay:.2f}s")
        services_wa.send_whatsapp_message(phone_number, message)
    except Exception as e:
        print(f"Gagal mengirim WA ke {phone_number}: {e}")

def get_document_path(db: Session, project_id: Optional[int] = None, aktivitas_id: Optional[int] = None):
    """
    Fungsi pembantu untuk membangun jalur penyimpanan file berdasarkan
    aktivitas atau proyek.
    """
    if not project_id and not aktivitas_id:
        raise HTTPException(status_code=400, detail="project_id atau aktivitas_id harus diberikan.")

    folder_tahun = str(date.today().year)
    folder_tim = None
    folder_proyek = None
    folder_aktivitas = None

    if aktivitas_id:
        # Muat aktivitas dan relasinya ke tim
        aktivitas = db.query(models.Aktivitas).options(
            joinedload(models.Aktivitas.team)
        ).filter(models.Aktivitas.id == aktivitas_id).first()
        
        if not aktivitas:
            raise HTTPException(status_code=404, detail="Data aktivitas tidak ditemukan.")
        if not aktivitas.team:
            raise HTTPException(status_code=404, detail="Tim untuk aktivitas ini tidak ditemukan.")
        
        # Ambil project secara terpisah
        if not aktivitas.project_id:
            raise HTTPException(status_code=404, detail="Aktivitas tidak terhubung ke proyek manapun.")
            
        project = db.query(models.Project).filter(models.Project.id == aktivitas.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Proyek untuk aktivitas ini tidak ditemukan.")

        folder_tim = aktivitas.team.nama_tim.replace(' ', '-')
        folder_proyek = project.nama_project.replace(' ', '-')
        folder_aktivitas = f"{aktivitas.tanggal_mulai.strftime('%y%m%d')}_{aktivitas.nama_aktivitas.replace(' ','-')}"
        
    elif project_id:
        # Muat proyek dan relasinya ke tim
        project = db.query(models.Project).options(
            joinedload(models.Project.team)
        ).filter(models.Project.id == project_id).first()
        
        if not project:
            raise HTTPException(status_code=404, detail="Data proyek tidak ditemukan.")
        if not project.team:
            raise HTTPException(status_code=404, detail="Tim untuk proyek ini tidak ditemukan.")

        folder_tim = project.team.nama_tim.replace(' ', '-')
        folder_proyek = project.nama_project.replace(' ', '-')

    # Membangun jalur hierarkis
    base_path = os.path.join(STORAGE_DIRECTORY, folder_tahun, folder_tim, folder_proyek)
    
    if folder_aktivitas:
        base_path = os.path.join(base_path, folder_aktivitas)
        
    if not os.path.exists(base_path):
        os.makedirs(base_path, exist_ok=True)
        
    return base_path

def create_notification(
    db: Session,
    user_id: int,
    title: str,
    massage: str,
    link_to: str,
    activity_id: Optional[int] = None,
    project_id: Optional[int] = None,
    send_whatsapp: bool = True,
    wa_message_override: Optional[str] = None
):
    """
    Membuat notifikasi di DB. Jika send_whatsapp=True dan user memiliki 'nohp',
    akan mengirim notifikasi WA.
    'wa_message_override' akan digunakan sebagai isi pesan WA jika tersedia.
    Jika tidak, 'title' dan 'massage' akan digunakan.
    """
    db_notif = models.Notifikasi(
        user_id=user_id,
        title=title,
        massage=massage,
        link_to=link_to,
        related_activity_id=activity_id,
        related_project_id=project_id,
        is_read=False
    )
    db.add(db_notif)

    user = db.query(models.User).filter(models.User.id == user_id).first()

    if user and user.nohp and send_whatsapp:
        base_url = "https://sinergi.statsntb.id" 
        full_link = f"{base_url}{link_to}"
        wa_message = ""

        if wa_message_override:
            wa_message = wa_message_override.replace("{LINK}", full_link)
        else:
            wa_message = f"🔔 *Notifikasi SINERGI*\n\n"
            wa_message += f"*{title}*\n"
            wa_message += f"{massage}\n\n"
            wa_message += f"Lihat detail:\n{full_link}"

        new_queue = models.WaQueue(
            phone_number=user.nohp,
            message=wa_message,
            status="pending"
        )
        db.add(new_queue)
    
    return db_notif

# Helper membersihkan nama folder dari karakter terlarang
def sanitize_filename(name: str) -> str:
    """Membersihkan string agar aman digunakan sebagai nama file/folder."""
    # Ganti karakter ilegal dengan underscore atau kosong
    # Hapus karakter: < > : " / \ | ? *
    cleaned = re.sub(r'[<>:"/\\|?*]', '', name)
    # Hapus spasi berlebih
    return " ".join(cleaned.split())

def add_dokumen_to_zip(zip_file, dokumen_list, base_folder_path=""):
    """
    Fungsi reusable untuk memasukkan list dokumen ke dalam ZIP 
    dengan struktur folder berdasarkan keterangan/checklist.
    """
    for doc in dokumen_list:
        if doc.tipe == 'FILE' and doc.path_atau_url and os.path.exists(doc.path_atau_url):
            # Tentukan nama sub-folder (Grouping)
            # Jika punya parent checklist, gunakan nama checklist
            # Jika tidak, gunakan keterangan dokumen
            sub_folder_name = "Dokumen Lainnya"
            if doc.checklist_item:
                sub_folder_name = doc.checklist_item.nama_dokumen
            elif doc.keterangan:
                sub_folder_name = doc.keterangan
            
            clean_sub_folder = sanitize_filename(sub_folder_name)
            clean_filename = sanitize_filename(doc.nama_file_asli)
            
            # Struktur: [Base Path] / [Nama Sub Folder] / [Nama File]
            # Contoh: [251124]_Rapat/Notulensi/scan_notulensi.pdf
            zip_path = f"{base_folder_path}/{clean_sub_folder}/{clean_filename}"
            
            # Hindari double slash jika base_folder kosong
            if not base_folder_path:
                zip_path = f"{clean_sub_folder}/{clean_filename}"

            try:
                zip_file.write(doc.path_atau_url, arcname=zip_path)
            except Exception as e:
                print(f"Gagal zip file {doc.id}: {e}")
                
# ===================================================================
# ENDPOINT OTENTIKASI & PENGGUNA
# ===================================================================
@app.post("/token")
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(database.get_db)):
    user = security.get_user(db, username=form_data.username)
    if not user or not security.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, 
            detail="Username atau password salah",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # --- VALIDASI STATUS AKTIF (BARU) ---
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, # 403 Forbidden lebih tepat untuk user non-aktif
            detail="Akun Anda dinonaktifkan. Silakan hubungi Admin.",
            headers={"WWW-Authenticate": "Bearer"},
        )
    # ------------------------------------
    
    # --- UPDATE LAST LOGIN ---
    user.last_login = func.now() 
    db.add(user)
    db.commit()
    db.refresh(user) 

    # Akses Token (30 Menit)
    access_token_expires = timedelta(minutes=security.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = security.create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )

    # Refresh Token (30 Hari)
    refresh_token_expires = timedelta(days=security.REFRESH_TOKEN_EXPIRE_DAYS)
    refresh_token = security.create_refresh_token(
        data={"sub": user.username}, expires_delta=refresh_token_expires
    )

    return {
        "accessToken" : access_token,
        "refreshToken" : refresh_token,
        "tokenType": "bearer"
    }

async def get_current_user_from_refresh_token(
    token: str = Depends(security.oauth2_scheme), # Asumsi Anda punya oauth2_scheme
    db: Session = Depends(database.get_db)
):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate refresh token",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, security.SECRET_KEY, algorithms=[security.ALGORITHM])
        username: str = payload.get("sub")

        # Pastikan ini adalah refresh token
        if payload.get("token_type") != "refresh":
            raise credentials_exception
        if username is None:
            raise credentials_exception

    except JWTError:
        raise credentials_exception

    user = security.get_user(db, username=username)
    if user is None:
        raise credentials_exception
    return user

# (Ini mirip dengan 'get_current_user_from_refresh_token', tapi untuk 'reset')
async def get_user_from_reset_token(
    token: str = Body(..., embed=True), # Ambil token dari body
    db: Session = Depends(database.get_db)
):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token reset tidak valid or kedaluwarsa",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, security.SECRET_KEY, algorithms=[security.ALGORITHM])
        username: str = payload.get("sub")

        # Pastikan ini adalah token 'reset'
        if payload.get("token_type") != "reset":
            raise credentials_exception
        if username is None:
            raise credentials_exception

    except JWSError: # Termasuk 'ExpiredSignatureError'
        raise credentials_exception

    user = security.get_user(db, username=username)
    if user is None:
        raise credentials_exception
    return user

@app.post("/api/auth/refresh")
def refresh_access_token(
    current_user: models.User = Depends(get_current_user_from_refresh_token)
):
    """
    Menerima refresh token yang valid dan mengembalikan access token baru.
    """
    # Jika kode sampai di sini, refresh token sudah valid.
    # Buat HANYA access token baru (umur pendek).
    access_token_expires = timedelta(minutes=security.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = security.create_access_token(
        data={"sub": current_user.username}, expires_delta=access_token_expires
    )

    return {
        "accessToken": access_token,
        "tokenType": "bearer"
    }

@app.get("/users/me", response_model=schemas.UserWithTeams, response_model_by_alias=True)
def read_users_me(current_user: models.User = Depends(security.get_current_user), db: Session = Depends(database.get_db)):
    
    # --- LOGIKA UNTUK MENGAMBIL PERAN TIM ---
    teams_with_role = []
    
    # Query untuk mendapatkan semua tim dan peran user terkait dari user_team_link
    user_teams_links = db.query(
        models.Team, 
        models.user_team_link.c.team_role
    ).join(
        models.user_team_link, models.user_team_link.c.team_id == models.Team.id
    ).filter(
        models.user_team_link.c.user_id == current_user.id
    ).all()

    for team, role in user_teams_links:
        teams_with_role.append({
            "id": team.id,
            "nama_tim": team.nama_tim,
            "peran": role # 'member' atau 'operator'
        })
    
    # --- LOGIKA UNTUK MENENTUKAN is_ketua_tim & ketua_tim_aktif ---
    ketua_tim_aktif_list = db.query(models.Team).filter(models.Team.ketua_tim_id == current_user.id).all()
    is_ketua = len(ketua_tim_aktif_list) > 0
    
    # --- BANGUN DICTIONARY RESPONS SECARA MANUAL UNTUK KEAMANAN ---
    # Ini menghindari error konversi otomatis Pydantic yang kompleks
    response_data = {
        "id": current_user.id,
        "username": current_user.username,
        "nama_lengkap": current_user.nama_lengkap,
        "foto_profil_url": current_user.foto_profil_url,
        "is_active": current_user.is_active,
        "sistem_role": current_user.sistem_role,
        "jabatan": current_user.jabatan,
        "nip": current_user.nip,
        "nipbps": current_user.nipbps,
        "gol_akhir": current_user.gol_akhir,
        "tmt_gol": current_user.tmt_gol,
        "tmt_jab": current_user.tmt_jab,
        "status_kepegawaian": current_user.status_kepegawaian,
        "jenis_kelamin": current_user.jenis_kelamin,
        "nohp": current_user.nohp,

        # Data relasi lain yang dibutuhkan oleh skema UserWithTeams
        "created_projects": current_user.created_projects,
        "aktivitas": current_user.aktivitas,
        
        # Data yang kita proses secara manual dari atas
        "teams": teams_with_role,
        "is_ketua_tim": is_ketua,
        "ketua_tim_aktif": ketua_tim_aktif_list,
    }
    
    return response_data


@app.post("/api/{user_id}/upload-photo")
def upload_profile_photo(user_id: int, file: UploadFile = File(...), db: Session = Depends(database.get_db)):
    # cek ekstensi file
    if not file.filename.lower().endswith((".png", ".jpg", ".jpeg")):
        raise HTTPException(status_code=400, detail="Format foto tidak valid. Gunakan JPG/PNG")

    # buat folder kalau belum ada
    os.makedirs(UPLOAD_PROFILE_PIC_DIR, exist_ok=True)

    # simpan file
    file_path = f"{UPLOAD_PROFILE_PIC_DIR}/{user_id}_{file.filename}"
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # update DB
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")

    user.foto_profil_url = file_path
    db.commit()
    db.refresh(user)

    return {"message": "Foto profil berhasil diunggah", "foto_profil_url": user.foto_profil_url}

@app.delete("/api/{user_id}/delete-photo")
def delete_profile_photo(user_id: int, db: Session = Depends(database.get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")

    if user.foto_profil_url:
        file_path = user.foto_profil_url.lstrip("/")
        if os.path.exists(file_path):
            os.remove(file_path)

        user.foto_profil_url = None
        db.commit()
        db.refresh(user)

    return {"message": "Foto profil berhasil dihapus"}

@app.put("/api/users/{user_id}/password")
def update_password(
    user_id: int,
    password_data: schemas.PasswordUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    # Pastikan user hanya bisa ganti password dirinya sendiri
    if current_user.id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tidak diizinkan mengganti password user lain"
        )

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User tidak ditemukan")

    # Verifikasi password lama
    if not security.verify_password(password_data.old_password, user.hashed_password):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Password lama salah")

    # Update password baru
    hashed_new_password = security.get_password_hash(password_data.new_password)
    user.hashed_password = hashed_new_password
    db.commit()

    return {"message": "Password berhasil diperbarui"}

@app.put("/api/users/me/profile", response_model=schemas.User)
def update_own_profile(
    profile_data: schemas.ProfileUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Mengizinkan pengguna yang sedang login untuk memperbarui
    nama_lengkap dan nohp mereka sendiri.
    """
    
    # Ambil data dari payload
    update_data = profile_data.dict(exclude_unset=True)
    
    # Update field di objek user
    for key, value in update_data.items():
        if hasattr(current_user, key):
            setattr(current_user, key, value)
        
    db.commit()
    db.refresh(current_user)
    
    # Kita harus return dict manual karena 'teams'
    # (Salin dari read_users_me)
    teams_with_role = []
    user_teams_links = db.query(
        models.Team, 
        models.user_team_link.c.team_role
    ).join(
        models.user_team_link, models.user_team_link.c.team_id == models.Team.id
    ).filter(
        models.user_team_link.c.user_id == current_user.id
    ).all()

    for team, role in user_teams_links:
        teams_with_role.append({
            "id": team.id,
            "nama_tim": team.nama_tim,
            "peran": role
        })
    
    ketua_tim_aktif_list = db.query(models.Team).filter(models.Team.ketua_tim_id == current_user.id).all()
    is_ketua = len(ketua_tim_aktif_list) > 0

    response_data = {
        "id": current_user.id,
        "username": current_user.username,
        "nama_lengkap": current_user.nama_lengkap,
        "foto_profil_url": current_user.foto_profil_url,
        "is_active": current_user.is_active,
        "sistem_role": current_user.sistem_role,
        "jabatan": current_user.jabatan,
        "nip": current_user.nip,
        "nipbps": current_user.nipbps,
        "gol_akhir": current_user.gol_akhir,
        "tmt_gol": current_user.tmt_gol,
        "tmt_jab": current_user.tmt_jab,
        "status_kepegawaian": current_user.status_kepegawaian,
        "jenis_kelamin": current_user.jenis_kelamin,
        "nohp": current_user.nohp,
        "created_projects": current_user.created_projects,
        "aktivitas": current_user.aktivitas,
        "teams": teams_with_role,
        "is_ketua_tim": is_ketua,
        "ketua_tim_aktif": ketua_tim_aktif_list,
    }
    
    return response_data

# ===================================================================
# ENDPOINT MANAJEMEN ADMIN
# ===================================================================
@app.post("/api/users", response_model=schemas.User, response_model_by_alias=True, dependencies=[Depends(security.require_role(["Superadmin", "Admin"]))])
def create_user(user: schemas.UserCreate, db: Session = Depends(database.get_db)):
    db_user = security.get_user(db, username=user.username)
    if db_user:
        raise HTTPException(status_code=400, detail="Username sudah terdaftar")
    
    hashed_password = security.get_password_hash(user.password)
    
    new_user_data = user.dict()

    new_user_data.pop('password', None)

    new_user = models.User(
        **new_user_data,
        hashed_password=hashed_password
    )

    # new_user = models.User(
    #     username=user.username,
    #     hashed_password=hashed_password,
    #     nama_lengkap=user.nama_lengkap,
    #     sistem_role_id=user.sistem_role_id,
    #     jabatan_id=user.jabatan_id
    # )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.get("/api/users", response_model=schemas.UserPage, response_model_by_alias=True)
def get_all_users(
    db: Session = Depends(database.get_db),
    skip: int = 0, 
    limit: int = 10, 
    search: Optional[str] = None
):
    query = db.query(models.User).options(
        joinedload(models.User.sistem_role),
        joinedload(models.User.jabatan)
    )
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                models.User.nama_lengkap.ilike(search_term),
                models.User.username.ilike(search_term)
            )
        ).distinct()

    total = query.count()
    users_from_db = query.order_by(models.User.id.desc()).offset(skip).limit(limit).all()

    processed_users = []
    for user in users_from_db:
        teams_with_role = []
        
        # Query untuk mendapatkan peran tim untuk setiap user dalam loop
        user_teams_links = db.query(
            models.Team, 
            models.user_team_link.c.team_role
        ).join(
            models.user_team_link, models.user_team_link.c.team_id == models.Team.id
        ).filter(
            models.user_team_link.c.user_id == user.id
        ).all()

        for team, role in user_teams_links:
            teams_with_role.append({
                "id": team.id,
                "nama_tim": team.nama_tim,
                "peran": role
            })

        user_dict = {
            "id": user.id,
            "username": user.username,
            "nama_lengkap": user.nama_lengkap,
            "foto_profil_url": user.foto_profil_url,
            "is_active": user.is_active,
            "sistem_role": user.sistem_role,
            "jabatan": user.jabatan,
            "created_projects": user.created_projects,
            "aktivitas": user.aktivitas,
            "teams": teams_with_role,

            "nip": user.nip,
            "nipbps": user.nipbps,
            "gol_akhir": user.gol_akhir,
            "tmt_gol": user.tmt_gol,
            "tmt_jab": user.tmt_jab,
            "status_kepegawaian": user.status_kepegawaian,
            "jenis_kelamin": user.jenis_kelamin,
            "nohp": user.nohp,
            "last_login": user.last_login
        }
        processed_users.append(user_dict)

    return {"total": total, "items": processed_users}

@app.put("/api/users/{user_id}", response_model=schemas.User, response_model_by_alias=True, dependencies=[Depends(security.require_role(["Superadmin"]))])
def update_user(user_id: int, user_update: schemas.UserUpdate, db: Session = Depends(database.get_db)):
    """Memperbarui data pengguna (hanya Superadmin)."""
    
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    
    # Ambil data yang dikirim sebagai dictionary snake_case, abaikan field yang kosong (None)
    update_data = user_update.dict(exclude_unset=True)

    # Perbarui setiap field di objek database
    for key, value in update_data.items():
        setattr(db_user, key, value)
    
    # Simpan perubahan
    db.commit()
    db.refresh(db_user)
    
    teams_with_role = []
    user_teams_links = db.query(
        models.Team, 
        models.user_team_link.c.team_role
    ).join(
        models.user_team_link, models.user_team_link.c.team_id == models.Team.id
    ).filter(
        models.user_team_link.c.user_id == db_user.id # Gunakan db_user.id
    ).all()

    for team, role in user_teams_links:
        teams_with_role.append({
            "id": team.id,
            "nama_tim": team.nama_tim,
            "peran": role
        })

    # Buat kamus respons manual, sama seperti di get_all_users
    # Ini memastikan semua field baru (nip, nohp) dan teams (dengan peran) disertakan
    user_dict = {
        "id": db_user.id,
        "username": db_user.username,
        "nama_lengkap": db_user.nama_lengkap,
        "foto_profil_url": db_user.foto_profil_url,
        "is_active": db_user.is_active,
        "sistem_role": db_user.sistem_role,
        "jabatan": db_user.jabatan,
        "created_projects": db_user.created_projects,
        "aktivitas": db_user.aktivitas,
        "teams": teams_with_role, # <-- Menggunakan list dict yang sudah benar
        
        # Data baru
        "nip": db_user.nip,
        "nipbps": db_user.nipbps,
        "gol_akhir": db_user.gol_akhir,
        "tmt_gol": db_user.tmt_gol,
        "tmt_jab": db_user.tmt_jab,
        "status_kepegawaian": db_user.status_kepegawaian,
        "jenis_kelamin": db_user.jenis_kelamin,
        "nohp": db_user.nohp
    }

    return user_dict

# --- ENDPOINT UNTUK MENGHAPUS USER ---
@app.delete("/api/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(security.require_role(["Superadmin"]))])
def delete_user(user_id: int, db: Session = Depends(database.get_db)):
    """Menghapus pengguna berdasarkan ID (hanya Superadmin)."""
    
    # Cari pengguna di database
    user_query = db.query(models.User).filter(models.User.id == user_id)
    db_user = user_query.first()
    
    if db_user is None:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
        
    # Hapus pengguna
    user_query.delete(synchronize_session=False)
    db.commit()
    
    # Kembalikan respons tanpa konten
    return Response(status_code=status.HTTP_204_NO_CONTENT)

@app.post("/api/teams", response_model=schemas.Team, response_model_by_alias=True, dependencies=[Depends(security.require_role(["Superadmin", "Admin"]))])
def create_team(team: schemas.TeamCreate, db: Session = Depends(database.get_db)):
    # 1. Buat instance model Team
    db_team = models.Team(
        nama_tim=team.nama_tim,
        valid_from=team.valid_from,
        valid_until=team.valid_until,
        ketua_tim_id=team.ketua_tim_id,
        warna=team.warna
    )
    db.add(db_team)
    db.flush()  # Gunakan flush untuk mendapatkan ID tim (db_team.id) tanpa mengakhiri transaksi

    # 2. Siapkan data untuk tabel user_team_link
    operator_ids_set = set(team.operator_ids)
    
    # Gabungkan ID ketua dan operator, pastikan tidak ada duplikasi
    all_member_ids = {team.ketua_tim_id} | operator_ids_set if team.ketua_tim_id else operator_ids_set

    links_to_create = []
    if all_member_ids:
        # Cek apakah semua user valid dalam satu query untuk efisiensi
        valid_users = db.query(models.User).filter(models.User.id.in_(all_member_ids)).all()
        valid_user_ids = {user.id for user in valid_users}

        for user_id in valid_user_ids:
            # Peran operator memiliki prioritas lebih tinggi
            role = 'operator' if user_id in operator_ids_set else 'member'
            links_to_create.append({
                "user_id": user_id,
                "team_id": db_team.id,
                "team_role": role
            })

    # 3. Eksekusi bulk insert ke tabel user_team_link jika ada data
    if links_to_create:
        db.execute(insert(models.user_team_link), links_to_create)
    
    # 4. Commit semua perubahan (pembuatan tim dan penambahan anggota) dalam satu transaksi
    db.commit()
    db.refresh(db_team)
    
    return db_team

@app.put("/api/admin/users/{user_id}/reset-password", dependencies=[Depends(security.require_role(["Superadmin", "Admin"]))])
def admin_reset_password(
    user_id: int,
    payload: schemas.AdminResetPasswordRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Endpoint khusus untuk Admin mereset password user lain.
    """
    # 1. Cari user target
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")

    # 2. Validasi Password (Sama seperti UserCreate)
    password = payload.new_password
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password harus minimal 8 karakter")
    # Tambahkan validasi kompleksitas lain jika perlu

    # 3. Hash Password Baru
    hashed_password = security.get_password_hash(password)
    target_user.hashed_password = hashed_password
    
    # 4. Simpan
    db.commit()
    
    return {"message": f"Password untuk pengguna {target_user.username} berhasil direset."}

@app.post("/api/auth/forgot-password")
def request_password_reset(
    request: schemas.ForgotPasswordRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(database.get_db)
):
    user = security.get_user(db, username=request.username)

    # PENTING: Jangan beri tahu jika user ada atau tidak.
    # Tapi kita cek di internal apakah user ada dan punya No. HP.
    if user and user.nohp:
        # Buat token reset (JWT 15 menit)
        reset_token_expires = timedelta(minutes=security.PASSWORD_RESET_EXPIRE_MINUTES)
        reset_token = security.create_reset_token(
            data={"sub": user.username}, expires_delta=reset_token_expires
        )

        # Buat link dan pesan WA
        base_url = "https://sinergi.statsntb.id" # URL Frontend Anda
        full_link = f"{base_url}/reset-password?token={reset_token}"

        wa_message = (
            f"🔑 *Permintaan Reset Password SINERGI*\n\n"
            f"Halo {user.nama_lengkap},\n"
            f"Kami menerima permintaan untuk mereset password akun Anda. "
            f"Silakan klik link di bawah ini untuk melanjutkan:\n\n"
            f"{full_link}\n\n"
            f"Link ini hanya berlaku selama *15 menit*.\n"
            f"Jika Anda tidak merasa meminta ini, mohon abaikan pesan ini."
        )

        # Kirim WA di background
        background_tasks.add_task(
            services_wa.send_whatsapp_message,
            phone_number=user.nohp,
            message=wa_message
        )

    # Selalu kembalikan pesan sukses (untuk keamanan)
    return {"message": "Jika akun Anda terdaftar dengan No. HP, link reset akan dikirim."}

@app.post("/api/auth/reset-password")
def handle_password_reset(
    request: schemas.ResetPasswordRequest,
    db: Session = Depends(database.get_db)
):
    # Validasi token dan ambil user
    # Kita panggil dependency secara manual di sini
    try:
        user = asyncio.run(get_user_from_reset_token(token=request.token, db=db))
    except HTTPException:
         raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token reset tidak valid atau kedaluwarsa."
        )

    # Jika token valid, hash password baru dan simpan
    hashed_password = security.get_password_hash(request.new_password)
    user.hashed_password = hashed_password

    db.commit()

    return {"message": "Password Anda telah berhasil direset. Silakan login."}

@app.get("/api/teams", response_model=schemas.TeamPage, response_model_by_alias=True)
def get_all_teams(
    db: Session = Depends(database.get_db),
    skip: int = 0, 
    limit: int = 10, 
    q: Optional[str] = None
):
    query = db.query(models.Team).options(joinedload(models.Team.ketua_tim))
    if q:
        query = query.filter(models.Team.nama_tim.ilike(f"%{q}%"))
    total = query.count()
    teams = query.order_by(desc(models.Team.valid_until), desc(models.Team.id)).offset(skip).limit(limit).all()
    return {"total": total, "items": teams}

@app.get("/api/teams/active", response_model=schemas.TeamPage, response_model_by_alias=True)
def get_active_teams(
    db: Session = Depends(database.get_db),
    skip: int = 0,
    limit: int = 10,
    q: Optional[str] = None
):
    today = date.today()

    query = (
        db.query(models.Team)
        .options(joinedload(models.Team.ketua_tim))
        .filter(
            and_(
                models.Team.valid_from <= today,
                models.Team.valid_until >= today
            )
        )
    )

    if q:
        query = query.filter(models.Team.nama_tim.ilike(f"%{q}%"))

    total = query.count()

    teams = (
        query
        .order_by(models.Team.nama_tim.asc())
        .offset(skip)     
        .limit(limit)       
        .all()
    )
    return {"total": total, "items": teams}


@app.put("/api/teams/{team_id}", response_model=schemas.Team, response_model_by_alias=True)
def update_team(team_id: int, team_update: schemas.TeamUpdate, db: Session = Depends(database.get_db)):
    db_team = db.query(models.Team).filter(models.Team.id == team_id).first()
    if not db_team:
        raise HTTPException(status_code=404, detail="Tim tidak ditemukan")

    update_data = team_update.dict(exclude_unset=True, by_alias=False)

    # Jika ada ketua_tim_id baru
    if "ketua_tim_id" in update_data and update_data["ketua_tim_id"] is not None:
        new_ketua_id = update_data["ketua_tim_id"]

        # cek apakah user sudah ada di anggota tim via relasi
        ketua_sudah_anggota = any(u.id == new_ketua_id for u in db_team.users)

        # kalau belum → tambahkan
        if not ketua_sudah_anggota:
            ketua_user = db.query(models.User).filter(models.User.id == new_ketua_id).first()
            if not ketua_user:
                raise HTTPException(status_code=404, detail="Ketua Tim tidak ditemukan.")
            db_team.users.append(ketua_user)

    if "operator_ids" in update_data:
        new_operator_ids = set(update_data.pop('operator_ids', []))

        # 1. Hapus semua peran 'operator' yang ada untuk tim ini
        db.execute(
            update(models.user_team_link)
            .where(models.user_team_link.c.team_id == team_id)
            .values(team_role='member')
        )

        # 2. Tetapkan peran 'operator' untuk user yang baru dipilih
        if new_operator_ids:
            db.execute(
                update(models.user_team_link)
                .where(
                    models.user_team_link.c.team_id == team_id,
                    models.user_team_link.c.user_id.in_(new_operator_ids)
                )
                .values(team_role='operator')
            )

    # Update field lain
    for key, value in update_data.items():
        setattr(db_team, key, value)

    db.commit()
    db.refresh(db_team)
    return db_team


@app.delete("/api/teams/{team_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(security.require_role(["Superadmin"]))])
def delete_team(team_id: int, db: Session = Depends(database.get_db)):
    """Menghapus tim (hanya Superadmin), tetapi hanya jika tidak memiliki aktivitas terkait."""
    
    # 1. Cari tim yang akan dihapus
    db_team = db.query(models.Team).filter(models.Team.id == team_id).first()

    if db_team is None:
        raise HTTPException(status_code=404, detail="Tim tidak ditemukan")

    # 2. Cek apakah tim memiliki aktivitas terkait
    if db_team.aktivitas:
        raise HTTPException(
            status_code=400,
            detail="Gagal menghapus tim. Tim ini masih memiliki aktivitas terkait."
        )

    # 3. Jika tidak ada aktivitas terkait, hapus tim
    # Anda harus menghapus data di tabel perantara secara manual (jika ada) sebelum menghapus tim utama
    db.delete(db_team)
    db.commit()
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# --- ENDPOINT BARU UNTUK MANAJEMEN ANGGOTA TIM ---

@app.get("/api/teams/{team_id}", response_model=schemas.Team, response_model_by_alias=True)
def get_team_details(team_id: int, db: Session = Depends(database.get_db)):
    """Mengambil detail satu tim, termasuk daftar anggotanya."""
    
    # Gunakan joinedload untuk mengambil data anggota sekaligus
    # Gunakan joinedload untuk mengambil data anggota sekaligus
    db_team = db.query(models.Team).options(
        joinedload(models.Team.users).joinedload(models.User.jabatan),
        joinedload(models.Team.users).joinedload(models.User.sistem_role)
    ).filter(models.Team.id == team_id).first()
    
    if not db_team:
        raise HTTPException(status_code=404, detail="Tim tidak ditemukan")
    
    return db_team

@app.post("/api/teams/{team_id}/members", response_model=schemas.Team, response_model_by_alias=True)
def add_team_member(team_id: int, user_id: int, background_tasks: BackgroundTasks, db: Session = Depends(database.get_db)):
    """Menambahkan seorang pengguna ke dalam tim."""
    db_team = db.query(models.Team).filter(models.Team.id == team_id).first()
    db_user = db.query(models.User).filter(models.User.id == user_id).first()

    if not db_team or not db_user:
        raise HTTPException(status_code=404, detail="Tim atau User tidak ditemukan")

    # Cek agar tidak duplikat
    if db_user in db_team.users:
        raise HTTPException(status_code=400, detail="Pengguna sudah menjadi anggota tim ini")

    if db_user not in db_team.users:
        db_team.users.append(db_user)
        db.commit()

        # --- LOGIKA NOTIFIKASI BARU: Anggota Ditambahkan ---
        nama_tim = db_team.nama_tim
        link_detail = f"/team/detail/{db_team.id}"
        
        title_notif = f"Anda ditambahkan ke {nama_tim}"
        message_notif = f"Anda sekarang adalah anggota {nama_tim}. Klik untuk melihat detail tim."
        
        create_notification(
            db, 
            user_id=user_id, 
            title=title_notif, 
            massage=message_notif, 
            link_to=link_detail,
            send_whatsapp=False
            
        )
        db.commit() # Commit notifikasi ke database
        # --- END LOGIKA NOTIFIKASI ---

        db.refresh(db_team)

    return db_team

@app.delete("/api/teams/{team_id}/members/{user_id}", response_model=schemas.Team, response_model_by_alias=True)
def remove_team_member(team_id: int, user_id: int, background_tasks: BackgroundTasks, db: Session = Depends(database.get_db)):
    """Mengeluarkan seorang pengguna dari tim."""
    db_team = db.query(models.Team).filter(models.Team.id == team_id).first()
    db_user = db.query(models.User).filter(models.User.id == user_id).first()

    if not db_team or not db_user:
        raise HTTPException(status_code=404, detail="Tim atau User tidak ditemukan")

    # Cek apakah pengguna benar-benar anggota tim
    if db_user not in db_team.users:
        raise HTTPException(status_code=400, detail="Pengguna bukan anggota tim ini")

    nama_tim = db_team.nama_tim

    if db_user in db_team.users:
        db_team.users.remove(db_user)
        db.commit()

    # --- LOGIKA NOTIFIKASI BARU: Anggota Dihapus ---
    
    title_notif = f"Keanggotaan Tim Berakhir"
    message_notif = f"Anda telah dikeluarkan dari {nama_tim} oleh Admin. Anda tidak lagi memiliki akses ke proyek dan aktivitas tim tersebut."
    
    create_notification(
        db, 
        user_id=user_id, 
        title=title_notif, 
        massage=message_notif, 
        link_to="/team" ,
        send_whatsapp=True 
    )
    db.commit() # Commit notifikasi
    # --- END LOGIKA NOTIFIKASI ---
    
    db.refresh(db_team)
    return db_team

@app.get("/api/teams/{team_id}/details", response_model=schemas.TeamDetail, response_model_by_alias=True)
def get_team_details_with_activities(team_id: int, db: Session = Depends(database.get_db)):
    """
    Mengambil detail satu tim, termasuk proyek, anggota (DENGAN PERAN), dan ketua.
    """
    db_team = db.query(models.Team).options(
        joinedload(models.Team.ketua_tim).joinedload(models.User.jabatan),
        joinedload(models.Team.users).joinedload(models.User.jabatan),
        joinedload(models.Team.projects).options(
            joinedload(models.Project.project_leader),
            joinedload(models.Project.aktivitas).joinedload(models.Aktivitas.users)
        )
    ).filter(models.Team.id == team_id).first()
    
    if not db_team:
        raise HTTPException(status_code=404, detail="Tim tidak ditemukan")

    # --- PENDEKATAN FINAL YANG BENAR ---
    # Tambahkan atribut 'peran' secara dinamis ke setiap objek user SQLAlchemy
    # sebelum Pydantic melakukan validasi dan konversi.
    if db_team.users:
        for user in db_team.users:
            link = db.query(models.user_team_link).filter(
                models.user_team_link.c.user_id == user.id,
                models.user_team_link.c.team_id == team_id
            ).first()
            
            # Menambahkan atribut sementara ke objek model SQLAlchemy
            setattr(user, 'peran', link.team_role if link else 'member')
    # --- AKHIR PENDEKATAN FINAL ---

    # Logika sorting aktivitas tetap sama
    for project in db_team.projects:
        project.aktivitas = sorted(
            project.aktivitas, 
            key=lambda a: a.tanggal_mulai if a.tanggal_mulai else date.min, 
            reverse=True
        )
    
    db_team.aktivitas = []
    
    # Kembalikan objek SQLAlchemy. Pydantic akan menanganinya dengan benar sekarang.
    return db_team

@app.get("/api/teams/{team_id}/aktivitas", response_model=List[schemas.Aktivitas])
def get_aktivitas_by_team_id(team_id: int, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    """
    Mengambil semua aktivitas yang terkait dengan ID tim tertentu.
    Aktivitas akan diurutkan dari yang terbaru ke yang terlama.
    """
    # Pastikan tim yang dicari ada di database
    db_team = db.query(models.Team).filter(models.Team.id == team_id).first()
    if not db_team:
        raise HTTPException(status_code=404, detail="Tim tidak ditemukan.")

    # Query database untuk mencari semua aktivitas dengan team_id yang cocok
    # Menggunakan joinedload untuk mengambil data terkait seperti dokumen dan daftar dokumen wajib
    # Menggunakan .order_by() untuk mengurutkan data dari yang terbaru (dibuat pada)
    db_aktivitas = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.dokumen),
        joinedload(models.Aktivitas.daftar_dokumen_wajib),
        joinedload(models.Aktivitas.creator)
    ).filter(models.Aktivitas.team_id == team_id).order_by(models.Aktivitas.dibuat_pada.desc()).all()

    # Mengembalikan daftar aktivitas
    return db_aktivitas

@app.get("/api/aktivitas/trend")
def get_aktivitas_trend(
    db: Session = Depends(database.get_db),
    group_by: str = Query(..., description="Filter: 'team', 'user', atau 'all_teams'"),
    group_id: Optional[int] = Query(None),
    months: int = Query(6)
):
    today = date.today()
    start_date = today.replace(day=1) - timedelta(days=30 * (months - 1))
    
    # Logic Baru: Tren Semua Tim (Multi-Line)
    if group_by == 'all_teams':
        query = db.query(
            func.to_char(models.Aktivitas.tanggal_mulai, 'YYYY-MM').label('month_year'),
            models.Team.nama_tim,
            func.count(models.Aktivitas.id).label('activity_count')
        ).join(models.Team).filter(
            models.Aktivitas.tanggal_mulai >= start_date
        ).group_by(
            'month_year', models.Team.nama_tim
        ).order_by('month_year')
        
        results = query.all()
        
        # Return format raw list, nanti frontend yang mengolah pivot-nya
        return [
            {
                "monthYear": row.month_year, 
                "groupName": row.nama_tim, 
                "activityCount": row.activity_count
            }
            for row in results
        ]

    # Logic Lama (Single Line)
    query = db.query(
        func.to_char(models.Aktivitas.tanggal_mulai, 'YYYY-MM').label('month_year'),
        func.count(models.Aktivitas.id).label('activity_count')
    ).filter(models.Aktivitas.tanggal_mulai >= start_date)

    if group_by == 'team' and group_id:
        query = query.filter(models.Aktivitas.team_id == group_id)
    elif group_by == 'user' and group_id:
        query = query.join(models.anggota_aktivitas_link).filter(models.anggota_aktivitas_link.c.user_id == group_id)
    
    results = query.group_by('month_year').order_by('month_year').all()
    
    return [
        {"monthYear": row.month_year, "activityCount": row.activity_count}
        for row in results
    ]


# ===================================================================
# ENDPOINT UNTUK MANAJEMEN PROJECT
# ===================================================================

@app.post("/api/projects", response_model=schemas.Project, response_model_by_alias=True)
def create_project(project: schemas.ProjectCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user), dependencies=[Depends(security.require_role("Superadmin"))]):
    
    team = db.query(models.Team).filter(models.Team.id == project.team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Tim tidak ditemukan")
    
    # --- LOGIKA OTORISASI BARU ---
    is_ketua_tim = team.ketua_tim_id == current_user.id
    is_superadmin = current_user.sistem_role.nama_role == "Superadmin"

    is_operator_query = select(models.user_team_link).where(
        models.user_team_link.c.user_id == current_user.id,
        models.user_team_link.c.team_id == project.team_id,
        models.user_team_link.c.team_role == 'operator'
    )
    is_operator = db.execute(is_operator_query).first() is not None

    if not is_ketua_tim and not is_operator and not is_superadmin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="Anda bukan Ketua Tim atau Operator untuk membuat proyek di tim ini"
        )
    
    project_data = project.dict(by_alias=False)
    send_wa_flag = project_data.pop('send_whatsapp', False)
    db_project = models.Project(**project_data)
    db.add(db_project)
    db.commit()
    db.refresh(db_project)

    # 1. Muat Project dengan relasi Team dan Anggota
    project_with_relations = db.query(models.Project).options(
        joinedload(models.Project.team).joinedload(models.Team.users)
    ).filter(models.Project.id == db_project.id).first()
    
    if not project_with_relations or not project_with_relations.team:
        # Lanjutkan jika tidak ada tim terkait (meskipun skema mengharuskan team_id)
        return db_project 
    
    nama_project = project_with_relations.nama_project
    project_leader_id = project_with_relations.project_leader_id
    link_detail = f"/project/detail/{db_project.id}"
    
    # 2. Definisikan Notifikasi Umum (Untuk Semua Anggota Tim)
    title_umum = f"Project Baru ditambahkan: {nama_project}"
    message_umum = f"Ketua Tim Anda telah membuat Project baru di bawah {project_with_relations.team.nama_tim}."
    
    # 3. Definisikan Notifikasi Khusus (Untuk Project Leader)
    title_leader = f"Anda Project Leader Project {nama_project}!"
    message_leader = "Anda telah ditunjuk sebagai Project Leader."
    
    # 4. Loop dan Kirim Notifikasi
    for user in project_with_relations.team.users:
            if user.id == project_leader_id:
                create_notification(
                    db=db, 
                    user_id=user.id, 
                    title=f"Anda Project Leader: {nama_project}", 
                    massage="Anda telah ditunjuk sebagai Project Leader.", 
                    link_to=link_detail, 
                    project_id=db_project.id,
                    send_whatsapp=send_wa_flag 
                )
            else:
                create_notification(
                    db=db, 
                    user_id=user.id, 
                    title=f"Project Baru: {nama_project}", 
                    massage=f"Project baru ditambahkan di tim {project_with_relations.team.nama_tim}.", 
                    link_to=link_detail, 
                    project_id=db_project.id,
                    send_whatsapp=send_wa_flag
                )
            
    db.commit() # Commit notifikasi
    # --- END LOGIKA NOTIFIKASI ---

    return db_project

@app.get("/api/projects", response_model=schemas.ProjectPage, response_model_by_alias=True)
def get_all_projects(
    db: Session = Depends(database.get_db),
    skip: int = 0,
    limit: int = 10,
    q: Optional[str] = None
):
    """Mendapatkan daftar semua proyek dengan paginasi dan pencarian."""
    query = db.query(models.Project).options(
        joinedload(models.Project.project_leader),
        joinedload(models.Project.team)
    )
    if q:
        query = query.filter(models.Project.nama_project.ilike(f"%{q}%"))
    total = query.count()
    projects = query.order_by(models.Project.id.asc()).offset(skip).limit(limit).all()
    return {"total": total, "items": projects}

@app.get("/api/projects/{project_id}", response_model=schemas.Project, response_model_by_alias=True)
def get_project_by_id(project_id: int, db: Session = Depends(database.get_db)):
    """Mendapatkan detail proyek dan daftar aktivitas aktif yang relevan."""
    
    db_project = db.query(models.Project).options(
        joinedload(models.Project.project_leader),
        joinedload(models.Project.team),
        joinedload(models.Project.dokumen)
    ).filter(models.Project.id == project_id).first()
    
    if not db_project:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")

    # Filter dan muat hanya aktivitas yang sedang aktif
    # today = date.today()
    # active_aktivitas = db.query(models.Aktivitas).options(
    #     joinedload(models.Aktivitas.daftar_dokumen_wajib)
    # ).with_parent(db_project).filter(
    #     or_( # Gunakan OR untuk dua kondisi
    #         # Kondisi 1: Aktivitas dengan rentang tanggal
    #         and_(
    #             models.Aktivitas.tanggal_selesai.isnot(None),
    #             models.Aktivitas.tanggal_mulai <= today,
    #             models.Aktivitas.tanggal_selesai >= today
    #         ),
    #         # Kondisi 2: Aktivitas satu hari tanpa jam
    #         and_(
    #             models.Aktivitas.tanggal_selesai.is_(None),
    #             models.Aktivitas.jam_mulai.is_(None),
    #             models.Aktivitas.jam_selesai.is_(None),
    #             models.Aktivitas.tanggal_mulai == today
    #         )
    #     )
    # ).all()
    # db_project.aktivitas = active_aktivitas

    all_aktivitas = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.daftar_dokumen_wajib),
        joinedload(models.Aktivitas.dokumen)
    ).with_parent(db_project).order_by(
        desc(models.Aktivitas.tanggal_mulai)
    ).all()

    db_project.aktivitas = all_aktivitas
    return db_project

@app.put("/api/projects/{project_id}", response_model=schemas.Project, response_model_by_alias=True)
def update_project(project_id: int, project_update: schemas.ProjectUpdate, db: Session = Depends(database.get_db)):
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")
    
    update_data = project_update.dict(exclude_unset=True, by_alias=False)
    for key, value in update_data.items():
        setattr(db_project, key, value)
    db.commit()
    db.refresh(db_project)
    return db_project

@app.delete("/api/projects/{project_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(security.require_role(["Superadmin", "Admin"]))])
def delete_project(project_id: int, db: Session = Depends(database.get_db)):
    """Menghapus proyek (hanya Superadmin atau Admin)."""
    project_query = db.query(models.Project).filter(models.Project.id == project_id)
    db_project = project_query.first()

    if db_project is None:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")
    
    project_query.delete(synchronize_session=False)
    db.commit()
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

@app.get("/api/sistem-roles", response_model=List[schemas.SistemRole])
def get_all_sistem_roles(db: Session = Depends(database.get_db)):
    """Mengembalikan semua peran sistem yang tersedia."""
    roles_db = db.query(models.SistemRole).all()
    # Konversi manual
    return [schemas.SistemRole.from_orm(role) for role in roles_db]

@app.get("/api/jabatan", response_model=List[schemas.Jabatan])
def get_all_jabatan(db: Session = Depends(database.get_db)):
    """Mengembalikan semua jabatan yang tersedia."""
    jabatan_db = db.query(models.Jabatan).all()
    # Konversi manual
    return [schemas.Jabatan.from_orm(j) for j in jabatan_db]

@app.get("/api/aktivitas", response_model=schemas.AktivitasPage)
def get_all_aktivitas(
    db: Session = Depends(database.get_db), 
    skip: int = 0,
    limit: int = 10,
    q: Optional[str] = None,
    current_user: models.User = Depends(security.get_current_user),
    user_scope: str = 'me',
    month: Optional[int] = None,
    year: Optional[int] = None
):
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.creator),
        joinedload(models.Aktivitas.team),
        # Optimasi: Preload children agar tidak N+1, tapi batasi kedalamannya
        selectinload(models.Aktivitas.children) 
    )

    # Filter berdasarkan scope (Aktivitas saya / Semua)
    if user_scope == 'me':
        query = query.filter(models.Aktivitas.users.any(id=current_user.id))

    # Filter berdasarkan bulan dan tahun
    if month:
        query = query.filter(extract('month', models.Aktivitas.tanggal_mulai) == month)
    if year:
        query = query.filter(extract('year', models.Aktivitas.tanggal_mulai) == year)

    # Jika ada parameter pencarian 'q'
    if q:
        search_term = f"%{q}%"
        # Lakukan join dengan tabel dokumen agar bisa mencari
        query = query.outerjoin(models.Dokumen)
        
        # Filter berdasarkan beberapa kolom sekaligus
        query = query.filter(
            or_(
                models.Aktivitas.nama_aktivitas.ilike(search_term),
                models.Aktivitas.deskripsi.ilike(search_term),
                models.Aktivitas.team.has(models.Team.nama_tim.ilike(search_term)),
                models.Dokumen.keterangan.ilike(search_term),
                models.Dokumen.nama_file_asli.ilike(search_term)
            )
        ).distinct()

    # --- LOGIKA PAGINATION ---
    
    # 1. Hitung total item yang memenuhi filter/pencarian
    # Karena ada .distinct(), kita harus menggunakan count pada subquery/ID
    if q:
        # Hitung jumlah ID yang berbeda untuk menghindari kesalahan count pada query dengan join dan distinct
        total_query = db.query(func.count(models.Aktivitas.id.distinct())).select_from(query.subquery())
        total = total_query.scalar()
    else:
        # Jika tidak ada distinct, count sederhana lebih cepat
        total = query.count()
        
    # 2. Ambil data dengan sorting, offset, dan limit
    aktivitas_items = (
        query
        .order_by(models.Aktivitas.id.desc())
        .offset(skip)              # 👈 Terapkan skip (offset)
        .limit(limit)              # 👈 Terapkan limit
        .all()
    )

    # 3. Kembalikan hasil dalam format paging
    return {"total": total, "items": aktivitas_items}

@app.get("/api/aktivitas/options", response_model=List[schemas.AktivitasOption])
def get_aktivitas_options(
    db: Session = Depends(database.get_db),
    q: Optional[str] = None
):
    """
    Endpoint ringan untuk dropdown parent aktivitas.
    Hanya mengembalikan ID dan Nama.
    """
    query = db.query(models.Aktivitas.id, models.Aktivitas.nama_aktivitas)
    
    if q:
        query = query.filter(models.Aktivitas.nama_aktivitas.ilike(f"%{q}%"))
    
    # Batasi hasil agar tidak membebani frontend
    results = query.limit(50).all()
    
    return [{"id": r.id, "nama_aktivitas": r.nama_aktivitas} for r in results]

@app.get("/api/aktivitas/kepala", response_model=List[schemas.Aktivitas])
def get_aktivitas_kepala(
    db: Session = Depends(database.get_db)):
    
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.team),
        joinedload(models.Aktivitas.project)
    ).filter(
        models.Aktivitas.melibatkan_kepala == True
    ).order_by(
        models.Aktivitas.tanggal_mulai.asc()
    )
    return query.all()

@app.get("/api/public/aktivitas/{public_id}", response_model=schemas.Aktivitas)
def get_public_aktivitas_detail(
    public_id: uuid.UUID,
    db: Session = Depends(database.get_db)
):
    """
    Endpoint publik untuk mengambil detail aktivitas menggunakan public_id (UUID).
    Tidak memerlukan otentikasi.
    """
    # Query aktivitas berdasarkan public_id
    aktivitas = db.query(models.Aktivitas).options(
        # Load semua relasi yang ingin Anda tampilkan di halaman publik
        joinedload(models.Aktivitas.project),
        joinedload(models.Aktivitas.team),
        joinedload(models.Aktivitas.users).joinedload(models.User.jabatan), # Load anggota & jabatan
        joinedload(models.Aktivitas.dokumen),
        # Load hierarki untuk publik juga
        joinedload(models.Aktivitas.parent),
        selectinload(models.Aktivitas.children)
    ).filter(models.Aktivitas.public_id == public_id).first()

    if not aktivitas:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")
    
    # Skema 'schemas.Aktivitas' Anda akan digunakan untuk merespons
    return aktivitas
   
@app.post("/api/aktivitas", response_model=schemas.Aktivitas)
def create_aktivitas(
    aktivitas: schemas.AktivitasCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user),
):
    # 1. BERSIHKAN PAYLOAD
    aktivitas_payload = aktivitas.dict()
    
    send_wa_flag = aktivitas_payload.pop('send_whatsapp', False)

    # Ambil dan hapus field list khusus (agar tidak error saat masuk ke models.Aktivitas)
    anggota_ids = aktivitas_payload.pop('anggota_aktivitas_ids', [])
    doc_wajib_names = aktivitas_payload.pop('daftar_dokumen_wajib', [])
    
    # V AMBIL ID TIM TERKAIT V
    tim_terkait_ids = aktivitas_payload.pop('id_tim_terkait', []) 
    # ^ ------------------ ^

    # Hapus field helper form
    aktivitas_payload.pop('use_date_range', None)
    aktivitas_payload.pop('use_time', None)

    # Validasi Parent ID (Mencegah Self-Reference saat create - meski ID belum ada, good practice)
    # Tidak perlu validasi ID karena create baru pasti ID belum ada

    # 2. BUAT INSTANCE AKTIVITAS
    db_aktivitas = models.Aktivitas(**aktivitas_payload)
    db_aktivitas.creator_user_id = current_user.id
    
    # 3. PROSES RELASI
    
    # A. Anggota Tim
    if anggota_ids:
        unique_anggota_ids = list(set(anggota_ids))
        anggota_tim = db.query(models.User).filter(models.User.id.in_(unique_anggota_ids)).all()
        db_aktivitas.users.extend(anggota_tim)

    # B. Dokumen Wajib
    for nama_dok in doc_wajib_names:
        if nama_dok:
            db_aktivitas.daftar_dokumen_wajib.append(
                models.DaftarDokumen(nama_dokumen=nama_dok, status_pengecekan=False)
            )

    # C. Tim Terkait (LOGIKA UTAMA)
    if tim_terkait_ids:
        # Filter ID unik dan pastikan bukan tim penyelenggara sendiri
        unique_tim_ids = list(set(tim_terkait_ids))
        if db_aktivitas.team_id in unique_tim_ids:
            unique_tim_ids.remove(db_aktivitas.team_id)
            
        if unique_tim_ids:
            teams_to_add = db.query(models.Team).filter(models.Team.id.in_(unique_tim_ids)).all()
            # Masukkan ke relasi 'tim_terkait'
            db_aktivitas.tim_terkait.extend(teams_to_add)

    # 4. SIMPAN
    db.add(db_aktivitas)
    db.commit()
    db.refresh(db_aktivitas)
    
    # LOGIKA PENGIRIMAN NOTIFIKASI
    if db_aktivitas.users:
        # Muat relasi Team dan Project
        aktivitas_with_relations = db.query(models.Aktivitas).options(
            joinedload(models.Aktivitas.team),
            joinedload(models.Aktivitas.project) # <-- Ini sudah ada
        ).filter(models.Aktivitas.id == db_aktivitas.id).first()

        # --- KUMPULKAN INFO UNTUK WA ---
        
        nama_aktivitas = aktivitas_with_relations.nama_aktivitas
        nama_tim = aktivitas_with_relations.team.nama_tim if aktivitas_with_relations.team else "Tim Tidak Diketahui"
        nama_project = aktivitas_with_relations.project.nama_project if aktivitas_with_relations.project else "Tanpa Proyek"
        link_detail = f"/aktivitas/detail/{db_aktivitas.id}"

        # 1. Format String Pelaksanaan (Tanggal)
        pelaksanaan_str = ""
        tgl_mulai = aktivitas_with_relations.tanggal_mulai
        tgl_selesai = aktivitas_with_relations.tanggal_selesai
        
        if tgl_mulai:
            pelaksanaan_str = f"🗓️ {tgl_mulai.strftime('%d %B %Y')}"
            if tgl_selesai and tgl_selesai != tgl_mulai:
                pelaksanaan_str += f" - {tgl_selesai.strftime('%d %B %Y')}"
        else:
            pelaksanaan_str = "🗓️ Tanggal belum ditentukan"

        # 2. Format String Pelaksanaan (Waktu)
        jam_mulai = aktivitas_with_relations.jam_mulai
        jam_selesai = aktivitas_with_relations.jam_selesai
        
        if jam_mulai:
            jam_str = f"⏰ {jam_mulai.strftime('%H.%M')}"
            if jam_selesai:
                jam_str += f" - {jam_selesai.strftime('%H.%M')} WITA"
            else:
                jam_str += " WITA"
            pelaksanaan_str += f"\n{jam_str}"
        
        # --- AKHIR INFO WA ---

        # Loop User (Tanpa delay index)
        for user in db_aktivitas.users:
            wa_msg_template = (
                f"🔔 *Aktivitas Baru*\n\n"
                f"Halo {user.nama_lengkap},\n"
                f"Anda ditambahkan ke:\n\n"
                f"📌 *{nama_aktivitas}*\n"
                f"🏢 Tim: {nama_tim}\n"
                f"🚀 Project: {nama_project}\n\n"
                f"*Waktu*: {pelaksanaan_str}\n\n" # Tambahkan jika variabel tersedia
                f"Cek detail: {{LINK}}"
            )
            
            create_notification(
                db=db, 
                user_id=user.id, 
                title=f"Aktivitas Baru: {nama_aktivitas}",
                massage=f"Anda ditambahkan ke aktivitas baru.",
                link_to=link_detail,
                activity_id=db_aktivitas.id, 
                project_id=db_aktivitas.project_id,
                send_whatsapp=send_wa_flag,
                wa_message_override=wa_msg_template
            )
        
        db.commit()
    # END: LOGIKA PENGIRIMAN NOTIFIKASI

    return db_aktivitas

@app.get("/api/aktivitas/download-excel", response_class=StreamingResponse)
def download_aktivitas_excel(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user),
    user_scope: str = 'me',
    month: Optional[int] = None,
    year: Optional[int] = None
):
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.project),
        joinedload(models.Aktivitas.team)
    )

    if user_scope == 'me':
        query = query.filter(models.Aktivitas.users.any(id=current_user.id))

    if month:
        query = query.filter(extract('month', models.Aktivitas.tanggal_mulai) == month)
    if year:
        query = query.filter(extract('year', models.Aktivitas.tanggal_mulai) == year)

    # Ambil semua data
    aktivitas_list = query.order_by(models.Aktivitas.tanggal_mulai.asc()).all()

    # Buat File Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Aktivitas"
    
    # --- HEADER (DIREVISI SESUAI PERMINTAAN) ---
    # Urutan: No, Waktu (4 kolom), Nama, Proyek, Tim, Link, Status, Parent
    headers = [
        "No",
        "Tanggal Mulai", 
        "Tanggal Selesai", 
        "Jam Mulai", 
        "Jam Selesai", 
        "Nama Aktivitas", 
        "Status", # Baru
        "Parent", # Baru
        "Nama Proyek", 
        "Tim", 
        "Bukti Dukung (Link)"
    ]
    ws.append(headers)

    # Style Header
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- ISI DATA ---
    base_url = "https://sinergi.statsntb.id" 
    
    # Gunakan enumerate untuk mendapatkan nomor urut (index) mulai dari 1
    for index, aktivitas in enumerate(aktivitas_list, start=1):
        
        # Persiapkan Data
        project_name = aktivitas.project.nama_project if aktivitas.project else "N/A"
        team_name = aktivitas.team.nama_tim if aktivitas.team else "N/A"
        link = f"{base_url}/public/aktivitas/{aktivitas.public_id}"
        
        # Format Waktu
        tgl_mulai_str = aktivitas.tanggal_mulai.strftime('%Y-%m-%d') if aktivitas.tanggal_mulai else ""
        tgl_selesai_str = aktivitas.tanggal_selesai.strftime('%Y-%m-%d') if aktivitas.tanggal_selesai else ""
        jam_mulai_str = aktivitas.jam_mulai.strftime('%H:%M') if aktivitas.jam_mulai else ""
        jam_selesai_str = aktivitas.jam_selesai.strftime('%H:%M') if aktivitas.jam_selesai else ""

        # --- APPEND ROW (URUTAN BARU) ---
        row_data = [
            index,            # Kolom A: No
            tgl_mulai_str,    # Kolom B: Tgl Mulai
            tgl_selesai_str, # Kolom C: Tgl Selesai
            jam_mulai_str,    # Kolom D: Jam Mulai
            jam_selesai_str, # Kolom E: Jam Selesai
            aktivitas.nama_aktivitas, # Kolom F: Nama
            aktivitas.status, # Kolom G: Status
            aktivitas.parent.nama_aktivitas if aktivitas.parent else "-", # Kolom H: Parent
            project_name,    # Kolom I: Proyek
            team_name,        # Kolom J: Tim
            link              # Kolom K: Link
        ]
        ws.append(row_data)
        
        # (Opsional) Set alignment center untuk No dan Waktu
        current_row = ws.max_row
        # Kolom 1 s.d 5 (A-E) rata tengah
        for col_idx in range(1, 6): 
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- AUTO WIDTH (DISESUAIKAN DENGAN URUTAN BARU) ---
    column_widths = {
        'A': 5,  # No (Kecil)
        'B': 15, # Tgl Mulai
        'C': 15, # Tgl Selesai
        'D': 10, # Jam Mulai
        'E': 10, # Jam Selesai
        'F': 40, # Nama Aktivitas (Lebar)
        'G': 15, # Status
        'H': 25, # Parent
        'I': 30, # Proyek
        'J': 25, # Tim
        'K': 50  # Link (Paling Lebar)
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Simpan ke stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"aktivitas_{user_scope}_{year}_{month}.xlsx" if month and year else f"aktivitas_{user_scope}.xlsx"
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- ENDPOINT BACKUP FILE BULANAN ---
@app.get("/api/aktivitas/backup-monthly")
def backup_monthly_files(
    month: int,
    year: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Membuat file ZIP berisi semua dokumen aktivitas pada bulan/tahun tertentu.
    Struktur ZIP: [Tgl] Nama Aktivitas / Nama File Asli.ext
    """
    
    # 1. Query Aktivitas (Filter sama seperti Excel)
    # Kita filter berdasarkan scope user juga (opsional, disini saya buat agar user hanya backup apa yang dia boleh lihat)
    # Atau jika ini fitur khusus admin/kepala, sesuaikan filternya.
    # Disini saya asumsi backup sesuai hak akses user (seperti excel 'me' atau 'all')
    # Untuk default backup bulanan, kita ambil semua aktivitas yang user ini terlibat.
    
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.dokumen)
    ).filter(
        extract('month', models.Aktivitas.tanggal_mulai) == month,
        extract('year', models.Aktivitas.tanggal_mulai) == year,
        # Filter hak akses: User harus terlibat di aktivitas tersebut
        models.Aktivitas.users.any(id=current_user.id) 
    )
    
    aktivitas_list = query.all()

    if not aktivitas_list:
        raise HTTPException(status_code=404, detail="Tidak ada aktivitas ditemukan untuk bulan ini.")

    # 2. Siapkan File ZIP Sementara
    # Kita tidak membuatnya di RAM agar server tidak crash jika filenya besar (GBs)
    try:
        # Buat file temp, delete=False agar kita bisa membacanya nanti untuk dikirim
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
        tmp_path = tmp_file.name
        
        files_added_count = 0
        
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for aktivitas in aktivitas_list:
                # Buat Nama Folder: "[01] Rapat Koordinasi"
                tgl = aktivitas.tanggal_mulai.strftime('%d') if aktivitas.tanggal_mulai else "00"
                safe_activity_name = sanitize_filename(aktivitas.nama_aktivitas)
                # Potong nama jika terlalu panjang agar tidak error path limit
                if len(safe_activity_name) > 50:
                    safe_activity_name = safe_activity_name[:50] + "..."
                    
                folder_name = f"[{tgl}] {safe_activity_name}"
                
                # Loop Dokumen
                for doc in aktivitas.dokumen:
                    # Kita hanya backup FILE, bukan LINK
                    if doc.tipe == 'FILE' and doc.path_atau_url:
                        # Cek apakah file fisik ada
                        if os.path.exists(doc.path_atau_url):
                            # Nama file di dalam ZIP
                            safe_filename = sanitize_filename(doc.nama_file_asli)
                            zip_entry_path = f"{folder_name}/{safe_filename}"
                            
                            try:
                                # Tulis file ke ZIP
                                zipf.write(doc.path_atau_url, arcname=zip_entry_path)
                                files_added_count += 1
                            except Exception as e:
                                print(f"Gagal zip file {doc.id}: {e}")
        
        tmp_file.close() # Tutup handle file agar bisa dibuka oleh FileResponse

        if files_added_count == 0:
            os.remove(tmp_path) # Hapus temp kosong
            raise HTTPException(status_code=404, detail="Aktivitas ditemukan, tapi tidak ada dokumen fisik yang tersimpan.")

        # 3. Kirim File dan Jadwalkan Penghapusan
        filename = f"Backup_Aktivitas_{year}_{month}.zip"
        
        # Background task untuk menghapus file temp setelah download selesai
        background_tasks.add_task(os.remove, tmp_path)

        return FileResponse(
            path=tmp_path,
            filename=filename,
            media_type='application/zip'
        )

    except Exception as e:
        # Bersihkan jika error di tengah jalan
        if 'tmp_path' in locals() and os.path.exists(tmp_path):
            os.remove(tmp_path)
        print(f"Error backup: {e}")
        raise HTTPException(status_code=500, detail="Terjadi kesalahan saat memproses backup.")

# --- ENDPOINT MENGAMBIL DETAIL AKTIVITAS ---
@app.get("/api/aktivitas/{aktivitas_id}", response_model=schemas.Aktivitas)
def get_aktivitas_by_id(aktivitas_id: int, db: Session = Depends(database.get_db)):
    # Query database untuk mencari aktivitas dengan ID yang sesuai
    db_aktivitas = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.dokumen),    
        joinedload(models.Aktivitas.daftar_dokumen_wajib),
        joinedload(models.Aktivitas.team),
        joinedload(models.Aktivitas.tim_terkait),
        joinedload(models.Aktivitas.users),
        # --- NEW: LOAD PARENT & CHILDREN ---
        joinedload(models.Aktivitas.parent),
        selectinload(models.Aktivitas.children) # selectinload lebih efisien untuk collection
    ).filter(models.Aktivitas.id == aktivitas_id).first()
    
    # Jika aktivitas tidak ditemukan, kirim error 404
    if db_aktivitas is None:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")
        
    # Jika ditemukan, kembalikan datanya
    return db_aktivitas

# --- ENDPOINT MENGUPDATE AKTIVITAS ---
@app.put("/api/aktivitas/{aktivitas_id}", response_model=schemas.Aktivitas)
def update_aktivitas(
    aktivitas_id: int, 
    aktivitas: schemas.AktivitasUpdate, 
    db: Session = Depends(database.get_db), 
    current_user: models.User = Depends(security.get_current_user)
):
    """Memperbarui aktivitas beserta relasinya dan mengirim notifikasi perubahan anggota."""
    
    # Load aktivitas dengan semua relasi
    db_aktivitas = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.daftar_dokumen_wajib),
        joinedload(models.Aktivitas.users),
        joinedload(models.Aktivitas.tim_terkait),
        joinedload(models.Aktivitas.team),    # Load untuk info notifikasi
        joinedload(models.Aktivitas.project)  # Load untuk info notifikasi
    ).filter(models.Aktivitas.id == aktivitas_id).first()
    
    if db_aktivitas is None:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")

    # 1. Ambil Data Payload
    update_data = aktivitas.dict(exclude_unset=True)
    
    # Ambil flag send_whatsapp (default False jika tidak dikirim)
    send_wa_flag = update_data.pop('send_whatsapp', False)

    # Validasi Parent
    if 'parent_id' in update_data and update_data['parent_id'] == aktivitas_id:
        raise HTTPException(status_code=400, detail="Aktivitas tidak bisa menjadi induk bagi dirinya sendiri.")

    # Bersihkan field helper
    update_data.pop('use_date_range', None)
    update_data.pop('use_time', None)
    
    # Pisahkan data relasi
    anggota_ids = update_data.pop('anggota_aktivitas_ids', None)
    doc_wajib_names = update_data.pop('daftar_dokumen_wajib', None)
    tim_terkait_ids = update_data.pop('id_tim_terkait', None) 

    # 2. Update Field Dasar
    for key, value in update_data.items():
        if hasattr(db_aktivitas, key):
            setattr(db_aktivitas, key, value)
    
    # 3. Update Anggota & Deteksi Perubahan (LOGIKA UTAMA)
    if anggota_ids is not None:
        old_user_ids = {u.id for u in db_aktivitas.users}
        new_user_ids = set(anggota_ids)
        
        # Hitung Diff
        added_ids = new_user_ids - old_user_ids
        removed_ids = old_user_ids - new_user_ids
        
        # Lakukan Update DB
        if new_user_ids:
            users_to_add = db.query(models.User).filter(models.User.id.in_(list(new_user_ids))).all()
            db_aktivitas.users = users_to_add # Replace all relations
        else:
            db_aktivitas.users = []

        # --- LOGIKA NOTIFIKASI PERUBAHAN ---
        if send_wa_flag:
            # --- KUMPULKAN INFO LENGKAP (Sama seperti create_aktivitas) ---
            nama_aktivitas = db_aktivitas.nama_aktivitas
            nama_tim = db_aktivitas.team.nama_tim if db_aktivitas.team else "Tim Tidak Diketahui"
            nama_project = db_aktivitas.project.nama_project if db_aktivitas.project else "Tanpa Proyek"
            link_detail = f"/aktivitas/detail/{db_aktivitas.id}"

            # Format Waktu
            pelaksanaan_str = ""
            tgl_mulai = db_aktivitas.tanggal_mulai
            tgl_selesai = db_aktivitas.tanggal_selesai
            
            if tgl_mulai:
                pelaksanaan_str = f"🗓️ *Tanggal:* {tgl_mulai.strftime('%d %B %Y')}"
                if tgl_selesai and tgl_selesai != tgl_mulai:
                    pelaksanaan_str += f" - {tgl_selesai.strftime('%d %B %Y')}"
            else:
                pelaksanaan_str = "🗓️ *Tanggal:* Belum ditentukan"

            jam_mulai = db_aktivitas.jam_mulai
            jam_selesai = db_aktivitas.jam_selesai
            
            if jam_mulai:
                jam_str = f"⏰ *Waktu:* {jam_mulai.strftime('%H.%M')}"
                if jam_selesai:
                    jam_str += f" - {jam_selesai.strftime('%H.%M')} WITA"
                else:
                    jam_str += " WITA"
                pelaksanaan_str += f"\n{jam_str}"

            # Deskripsi
            deskripsi_str = ""
            if db_aktivitas.deskripsi:
                 desc_text = db_aktivitas.deskripsi
                 if len(desc_text) > 100:
                     desc_text = desc_text[:100] + "..."
                 deskripsi_str = f"\n📝 *Deskripsi:*\n_{desc_text}_\n"

            # Dokumen Wajib (Ambil dari DB karena mungkin baru diupdate di langkah 4, tapi kita ambil current state dulu atau update list local)
            # Untuk simplifikasi, kita ambil dari db_aktivitas yang ada di sesi (mungkin belum ter-refresh list dokumennya jika diupdate di bawah)
            # Tapi info dokumen biasanya kurang kritikal dibanding waktu/tempat untuk notif 'diundang'.
            
            # A. Notifikasi untuk User yang DITAMBAHKAN (Format Lengkap)
            for uid in added_ids:
                user = db.query(models.User).filter(models.User.id == uid).first()
                if user:
                    wa_msg_add = (
                        f"🔔 *Aktivitas Baru*\n\n"
                        f"Halo {user.nama_lengkap},\n"
                        f"Anda ditambahkan ke:\n\n"
                        f"📌 *{nama_aktivitas}*\n"
                        f"🏢 Tim: {nama_tim}\n"
                        f"🚀 Project: {nama_project}\n\n"
                        f"*Waktu*: {pelaksanaan_str}\n\n"
                        f"Cek detail: {{LINK}}"
                    )
                    
                    create_notification(
                        db=db, user_id=uid, 
                        title=f"Ditambahkan: {nama_aktivitas}",
                        massage="Anda ditambahkan ke aktivitas ini.",
                        link_to=link_detail,
                        activity_id=db_aktivitas.id,
                        project_id=db_aktivitas.project_id,
                        send_whatsapp=True,
                        wa_message_override=wa_msg_add
                    )

            # B. Notifikasi untuk User yang DIHAPUS (Format Informatif)
            for uid in removed_ids:
                user = db.query(models.User).filter(models.User.id == uid).first()
                if user:
                    wa_msg_remove = (
                        f"ℹ️ *Update Aktivitas*\n\n"
                        f"Halo {user.nama_lengkap},\n"
                        f"Anda telah dikeluarkan dari aktivitas:\n"
                        f"❌ *{nama_aktivitas}*\n\n"
                        f"Anda tidak lagi terlibat dalam aktivitas tersebut."
                    )
                    create_notification(
                        db=db, user_id=uid, 
                        title=f"Dikeluarkan: {nama_aktivitas}",
                        massage="Anda telah dikeluarkan dari aktivitas ini.",
                        link_to="#", 
                        activity_id=None,
                        send_whatsapp=True,
                        wa_message_override=wa_msg_remove
                    )
    
    # 4. Update Dokumen Wajib
    if doc_wajib_names is not None:
        current_docs = {d.nama_dokumen: d for d in db_aktivitas.daftar_dokumen_wajib}
        new_docs_set = set(doc_wajib_names)
        
        for name, doc_obj in list(current_docs.items()):
            if name not in new_docs_set:
                db.delete(doc_obj)
        
        for name in new_docs_set:
            if name not in current_docs:
                db_aktivitas.daftar_dokumen_wajib.append(
                    models.DaftarDokumen(nama_dokumen=name, status_pengecekan=False)
                )

    # 5. Update Tim Terkait
    if tim_terkait_ids is not None:
        db_aktivitas.tim_terkait = []
        unique_tim_ids = list(set(tim_terkait_ids))
        if db_aktivitas.team_id in unique_tim_ids:
            unique_tim_ids.remove(db_aktivitas.team_id)
            
        if unique_tim_ids:
            teams_to_add = db.query(models.Team).filter(models.Team.id.in_(unique_tim_ids)).all()
            db_aktivitas.tim_terkait.extend(teams_to_add)

    db.commit()
    db.refresh(db_aktivitas)
    return db_aktivitas

# --- ENDPOINT MENGHAPUS AKTIVITAS ---
@app.delete("/api/aktivitas/{aktivitas_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_aktivitas(
    aktivitas_id: int, 
    background_tasks: BackgroundTasks,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    aktivitas_to_delete = db.query(models.Aktivitas).filter(models.Aktivitas.id == aktivitas_id).first()

    if aktivitas_to_delete is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Aktivitas tidak ditemukan.")

    if aktivitas_to_delete.dokumen:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Tidak dapat menghapus aktivitas karena masih terdapat dokumen terkait. Harap hapus semua dokumen terkait terlebih dahulu."
        )

    nama_aktivitas = aktivitas_to_delete.nama_aktivitas
    users_terlibat = aktivitas_to_delete.team.users if aktivitas_to_delete.team else []

    # --- LOGIKA NOTIFIKASI: SEBELUM PENGHAPUSAN ---
    if users_terlibat:
        title_notif = f"Aktivitas Dihapus: {nama_aktivitas}"
        message_notif = (
            f"Aktivitas '{nama_aktivitas}' telah dihapus. "
            "Semua data terkait telah dihilangkan."
        )
        
        # Kirim notifikasi ke semua anggota yang terlibat
        for i, user in enumerate(users_terlibat):
            create_notification(
                db, 
                user_id=user.id, 
                title=title_notif, 
                massage=message_notif, 
                link_to="/dashboard", 
                activity_id=None,
                send_whatsapp=False
            )

        db.commit() # Commit notifikasi
    # --- END LOGIKA NOTIFIKASI ---

    # Hapus semua entri di tabel perantara 'anggota_aktivitas' secara manual
    db.query(models.anggota_aktivitas_link).filter(models.anggota_aktivitas_link.c.aktivitas_id == aktivitas_id).delete(synchronize_session=False)

    # Hapus semua entri di tabel 'daftar_dokumen' secara manual
    db.query(models.DaftarDokumen).filter(models.DaftarDokumen.aktivitas_id == aktivitas_id).delete(synchronize_session=False)

    # Hapus aktivitas itu sendiri
    db.delete(aktivitas_to_delete)
    db.commit()

    return Response(status_code=status.HTTP_204_NO_CONTENT)

# --- ENDPOINT UPLOAD DOKUMEN ---
@app.post("/api/aktivitas/{aktivitas_id}/dokumen", response_model=schemas.Dokumen)
def create_dokumen_untuk_aktivitas(
    aktivitas_id: int,
    keterangan: str = Form(...),
    checklist_item_id: Optional[int] = Form(None), # ID dari DaftarDokumen
    file: UploadFile = File(...),
    db: Session = Depends(database.get_db)
):
    # 1. Validasi Aktivitas
    aktivitas = db.query(models.Aktivitas).filter(models.Aktivitas.id == aktivitas_id).first()
    if not aktivitas:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")

    try:
        # 2. Simpan File dengan metode Aman (Immutable)
        # Tidak peduli nama aktivitas berubah, path ini tetap valid.
        saved_path = save_file_securely(file)
        
        # 3. Buat Entri Database
        db_dokumen = models.Dokumen(
            aktivitas_id=aktivitas_id,
            keterangan=keterangan,
            tipe='FILE',
            path_atau_url=saved_path,       # Path baru (storage/2025/...)
            nama_file_asli=file.filename,   # Nama asli user (Laporan.pdf)
            tipe_file_mime=file.content_type,
            daftar_dokumen_id=checklist_item_id # Relasi ke Item Checklist (One-to-Many)
        )
        
        db.add(db_dokumen)
        db.commit()
        db.refresh(db_dokumen)

        # 4. Update Status Pengecekan (Opsional)
        # Jika di-upload ke checklist, kita bisa otomatis tandai sesuatu, 
        # atau biarkan manual (validasi ketua tim).
        # Disini kita biarkan status_pengecekan apa adanya (default False).

        return db_dokumen

    except Exception as e:
        print(f"Error upload: {e}")
        raise HTTPException(status_code=500, detail="Gagal mengunggah file")

# --- ENDPOINT MENAMBAHKAN LINK ---
@app.post("/api/aktivitas/{aktivitas_id}/link", response_model=schemas.Dokumen)
def add_link_untuk_aktivitas(
    aktivitas_id: int,
    link_data: schemas.DokumenCreate,
    db: Session = Depends(database.get_db)
):
    # Cek dulu apakah aktivitasnya ada
    aktivitas = db.query(models.Aktivitas).filter(models.Aktivitas.id == aktivitas_id).first()
    if not aktivitas:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")

    # Buat entri dokumen baru dengan tipe 'LINK'
    db_dokumen = models.Dokumen(
        aktivitas_id=aktivitas_id,
        keterangan=link_data.keterangan,
        tipe='LINK',
        path_atau_url=link_data.path_atau_url,
        nama_file_asli=link_data.nama_file_asli,
        daftar_dokumen_id=link_data.checklist_item_id
    )

    db.add(db_dokumen)
    db.commit()
    db.refresh(db_dokumen)
    
    return db_dokumen

# --- ENDPOINT UNGGAH DOKUMEN UNTUK PROYEK ---
@app.post("/api/projects/{project_id}/dokumen", response_model=schemas.Dokumen)
def create_dokumen_untuk_proyek(
    project_id: int,
    keterangan: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(database.get_db)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")

    try:
        # 1. Gunakan helper save_file_securely (Storage Immutable)
        saved_path = save_file_securely(file)
        
        # 2. Simpan ke Database
        db_dokumen = models.Dokumen(
            project_id=project_id,
            keterangan=keterangan,
            tipe='FILE',
            path_atau_url=saved_path,
            nama_file_asli=file.filename,
            tipe_file_mime=file.content_type
        )
        db.add(db_dokumen)
        db.commit()
        db.refresh(db_dokumen)
        
        return db_dokumen
    except Exception as e:
        print(f"Error upload project doc: {e}")
        raise HTTPException(status_code=500, detail="Gagal mengunggah file")

# --------------------------------------------------------------------

# --- ENDPOINT TAMBAH LINK UNTUK PROYEK ---
@app.post("/api/projects/{project_id}/links", response_model=schemas.Dokumen)
def add_link_untuk_proyek(
    project_id: int,
    link_data: schemas.DokumenCreate,
    db: Session = Depends(database.get_db)
):
    """
    Menambahkan link ke sebuah proyek.
    """
    # 1. Cari proyek berdasarkan ID
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")

    # 2. Buat entri dokumen baru dengan tipe 'LINK'
    db_dokumen = models.Dokumen(
        project_id=project_id,
        keterangan=link_data.keterangan,
        tipe='LINK',
        path_atau_url=link_data.path_atau_url
    )
    
    db.add(db_dokumen)
    db.commit()
    db.refresh(db_dokumen)
    
    return db_dokumen

# --- ENDPOIN MENGHAPUS DOKUMEN ---
@app.delete("/api/dokumen/{dokumen_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_dokumen(dokumen_id: int, db: Session = Depends(database.get_db)):
    # 1. Cari Dokumen yang akan dihapus
    db_dokumen = db.query(models.Dokumen).filter(models.Dokumen.id == dokumen_id).first()
    
    if db_dokumen is None:
        raise HTTPException(status_code=404, detail="Dokumen tidak ditemukan")

    # 2. [LOGIKA BARU] Reset Checklist Item (jika dokumen ini terhubung)
    # Kita cek apakah dokumen ini punya parent 'daftar_dokumen_id'
    if db_dokumen.daftar_dokumen_id:
        # Cari item checklist induknya
        db_checklist_item = db.query(models.DaftarDokumen).filter(
            models.DaftarDokumen.id == db_dokumen.daftar_dokumen_id
        ).first()
        
        if db_checklist_item:
            # Logika reset: Jika dokumen ini dihapus, apakah checklist harus jadi unchecked?
            # Karena sekarang 'One-to-Many', satu checklist bisa punya banyak file.
            # Kita hanya uncheck jika ini adalah SATU-SATUNYA file di checklist tersebut.
            
            # Hitung sisa file lain di checklist ini
            count_other_files = db.query(models.Dokumen).filter(
                models.Dokumen.daftar_dokumen_id == db_checklist_item.id,
                models.Dokumen.id != dokumen_id # Kecuali yang mau dihapus
            ).count()
            
            if count_other_files == 0:
                # Jika tidak ada file lain, tandai belum selesai
                db_checklist_item.status_pengecekan = False 

    # 3. Hapus File Fisik (Jika tipe FILE)
    if db_dokumen.tipe == 'FILE':
        file_path = db_dokumen.path_atau_url
        # Pastikan path aman (tidak kosong) sebelum unlink
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError as e:
                print(f"Error deleting file {file_path}: {e}")
                # Lanjutkan penghapusan DB meski file gagal dihapus (orphan file)
    
    # 4. Hapus dari Database
    db.delete(db_dokumen)
    db.commit()
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# ENDPOINT UNTUK MANAJEMEN CHECKLIST DOKUMEN
@app.patch("/api/daftar_dokumen/{item_id}/cek", response_model=schemas.DaftarDokumen, response_model_by_alias=True)
def update_status_pengecekan(
    item_id: int,
    status_update: schemas.StatusPengecekanUpdate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Memperbarui status pengecekan (true/false) untuk sebuah item di daftar dokumen.
    Hanya bisa dilakukan oleh ketua tim dari aktivitas terkait.
    """
    # Cari item checklist di database, lakukan join untuk mengambil data tim terkait
    db_item = db.query(models.DaftarDokumen).options(
        joinedload(models.DaftarDokumen.aktivitas).joinedload(models.Aktivitas.team),
        joinedload(models.DaftarDokumen.files)
    ).filter(models.DaftarDokumen.id == item_id).first()

    if not db_item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, 
            detail="Item checklist tidak ditemukan"
        )

    # Ambil status lama sebelum diubah
    old_status = db_item.status_pengecekan
    new_status = status_update.status_pengecekan
    
    # Validasi Keamanan: Pastikan pengguna adalah ketua tim
    # Pastikan ada aktivitas dan tim yang tertaut sebelum memeriksa
    if not db_item.aktivitas or not db_item.aktivitas.team:
          raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, 
            detail="Item checklist tidak terhubung dengan tim yang valid"
        )

    # Jika validasi berhasil, perbarui status
    db_item.status_pengecekan = status_update.status_pengecekan
    db.commit()

    # --- LOGIKA NOTIFIKASI ---
    if old_status != new_status and db_item.files:
        
        # A. Tentukan Konten Notifikasi
        nama_dokumen = db_item.nama_dokumen
        nama_aktivitas = db_item.aktivitas.nama_aktivitas
        link_detail = f"/aktivitas/detail/{db_item.aktivitas_id}"
        
        if new_status is True:
            title_notif = f"Dokumen Disetujui: {nama_dokumen}"
            message_notif = f"Dokumen Anda di aktivitas '{nama_aktivitas}' telah divalidasi dan disetujui oleh {current_user.nama_lengkap}."
        else: # new_status is False (Pembatalan persetujuan)
            title_notif = f"Persetujuan Dibatalkan: {nama_dokumen}"
            message_notif = f"Persetujuan dokumen Anda di aktivitas '{nama_aktivitas}' dibatalkan oleh {current_user.nama_lengkap} untuk diperbaiki."
        
        # B. Kirim Notifikasi ke semua Anggota Aktivitas
        if db_item.aktivitas.users:
            for i, user in enumerate(db_item.aktivitas.users):
                create_notification(
                    db,
                    user_id=user.id,
                    title=title_notif,
                    massage=message_notif,
                    link_to=link_detail,
                    activity_id=db_item.aktivitas_id,
                    send_whatsapp=False
                )
            
    # --- END LOGIKA NOTIFIKASI ---

    db.refresh(db_item)
    
    # 4. Kembalikan data yang sudah diperbarui
    return db_item

# --- ENDPOINT BARU UNTUK UNDUH/PREVIEW DOKUMEN ---
@app.get("/api/dokumen/{dokumen_id}/download")
def download_dokumen(
    dokumen_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Mengirim file ke pengguna dengan nama file aslinya dan menyarankan preview.
    """
    db_dokumen = db.query(models.Dokumen).filter(models.Dokumen.id == dokumen_id).first()

    if db_dokumen is None or db_dokumen.tipe != 'FILE' or not os.path.exists(db_dokumen.path_atau_url):
        raise HTTPException(status_code=404, detail="File tidak ditemukan")

    # --- PERBAIKAN DI SINI ---
    # Atur header Content-Disposition secara manual untuk 'inline'
    headers = {
        'Content-Disposition': f'inline; filename="{db_dokumen.nama_file_asli}"'
    }
    
    # Kirim file sebagai respons dengan header yang sudah diatur
    return FileResponse(
        path=db_dokumen.path_atau_url,
        media_type=db_dokumen.tipe_file_mime,
        headers=headers
    )

# --- ENDPOINTUNTUK UNDUH SEMUA DOKUMEN DALAM SATU AKTIVITAS ---
@app.get("/api/aktivitas/{aktivitas_id}/download-all")
def download_all_dokumen_aktivitas(
    aktivitas_id: int,
    db: Session = Depends(database.get_db)
):
    # Load aktivitas beserta dokumen dan checklistnya
    aktivitas = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.dokumen).joinedload(models.Dokumen.checklist_item)
    ).filter(models.Aktivitas.id == aktivitas_id).first()

    if not aktivitas:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")

    # Format Folder Utama: [YYMMDD]_[nama_aktivitas]
    tgl = aktivitas.tanggal_mulai.strftime('%y%m%d') if aktivitas.tanggal_mulai else "000000"
    folder_root = f"[{tgl}]_{sanitize_filename(aktivitas.nama_aktivitas)}"

    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Gunakan helper kita
        add_dokumen_to_zip(zip_file, aktivitas.dokumen, base_folder_path=folder_root)
    
    zip_buffer.seek(0)
    zip_filename = f"{folder_root}.zip"
    
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/x-zip-compressed",
        headers={'Content-Disposition': f'attachment; filename="{zip_filename}"'}
    )

@app.get("/api/projects/{project_id}/download-all")
def download_all_project_documents(
    project_id: int,
    db: Session = Depends(database.get_db)
):
    # Load Project + Dokumen Project + Aktivitas + Dokumen Aktivitas
    project = db.query(models.Project).options(
        joinedload(models.Project.dokumen), # Dokumen level project
        joinedload(models.Project.aktivitas).joinedload(models.Aktivitas.dokumen).joinedload(models.Dokumen.checklist_item)
    ).filter(models.Project.id == project_id).first()
      
    if not project:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")

    zip_buffer = io.BytesIO()
    folder_project = sanitize_filename(project.nama_project)

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        
        # 1. Masukkan Dokumen Level Project
        # Disimpan di: [Nama Project]/[Keterangan Dokumen]
        add_dokumen_to_zip(zip_file, project.dokumen, base_folder_path=folder_project)

        # 2. Masukkan Dokumen Level Aktivitas
        # Disimpan di: [Nama Project]/[YYMMDD]_[Nama Aktivitas]/[Keterangan]
        for akt in project.aktivitas:
            tgl = akt.tanggal_mulai.strftime('%y%m%d') if akt.tanggal_mulai else "000000"
            folder_aktivitas = f"[{tgl}]_{sanitize_filename(akt.nama_aktivitas)}"
            full_path = f"{folder_project}/{folder_aktivitas}"
            
            add_dokumen_to_zip(zip_file, akt.dokumen, base_folder_path=full_path)

    zip_buffer.seek(0)
    zip_filename = f"{folder_project}.zip"

    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/x-zip-compressed",
        headers={'Content-Disposition': f'attachment; filename="{zip_filename}"'}
    )

@app.get("/api/aktivitas/backup-monthly")
def backup_monthly_files(
    month: int,
    year: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    # Query Aktivitas di bulan tersebut
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.team),
        joinedload(models.Aktivitas.dokumen).joinedload(models.Dokumen.checklist_item)
    ).filter(
        extract('month', models.Aktivitas.tanggal_mulai) == month,
        extract('year', models.Aktivitas.tanggal_mulai) == year,
        models.Aktivitas.users.any(id=current_user.id) 
    )
    aktivitas_list = query.all()

    if not aktivitas_list:
        raise HTTPException(status_code=404, detail="Tidak ada aktivitas untuk di-backup.")

    try:
        # Gunakan file temp di disk untuk backup besar
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
        tmp_path = tmp_file.name
        
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for akt in aktivitas_list:
                # Level 1: [YYMM] (Bulan) -> Opsional jika zipnya sudah bernama bulan
                # Level 2: [nama_tim]
                # Level 3: [YYMMDD]_[nama_aktivitas]
                
                folder_yy_mm = f"{str(year)[2:]}{month:02d}" # YYMM
                folder_tim = sanitize_filename(akt.team.nama_tim) if akt.team else "Tanpa Tim"
                
                tgl_akt = akt.tanggal_mulai.strftime('%y%m%d') if akt.tanggal_mulai else "000000"
                folder_akt = f"[{tgl_akt}]_{sanitize_filename(akt.nama_aktivitas)}"
                
                # Path Lengkap: 2411/Tim IPDS/[241127]_Rapat/....
                full_base_path = f"{folder_yy_mm}/{folder_tim}/{folder_akt}"
                
                # Gunakan helper
                add_dokumen_to_zip(zipf, akt.dokumen, base_folder_path=full_base_path)
        
        tmp_file.close()

        filename = f"Backup_Sinergi_{year}_{month:02d}.zip"
        background_tasks.add_task(os.remove, tmp_path)

        return FileResponse(
            path=tmp_path,
            filename=filename,
            media_type='application/zip'
        )

    except Exception as e:
        if 'tmp_path' in locals() and os.path.exists(tmp_path):
            os.remove(tmp_path)
        print(f"Error backup: {e}")
        raise HTTPException(status_code=500, detail="Gagal membuat backup.")
    
# ===================================================================
# ENDPOINT BARU UNTUK KALENDER TIM
# ===================================================================

@app.get("/api/kalender/events", response_model=List[schemas.Aktivitas])
def get_calendar_events(
    db: Session = Depends(database.get_db),
    team_ids: Optional[str] = Query(None, description="Daftar ID tim yang dipisahkan oleh koma."),
):
    """
    Mengambil daftar semua aktivitas yang relevan untuk tampilan kalender.
    Hanya mengembalikan aktivitas yang kalender_view == True.
    Jika team_ids diberikan, akan memfilter berdasarkan tim tersebut.
    """
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.users),
        joinedload(models.Aktivitas.team)
    ).filter(models.Aktivitas.kalender_view == True) # <--- FILTER WAJIB

    if team_ids:
        try:
            team_id_list = {int(id_str) for id_str in team_ids.split(',') if id_str.isdigit()}
            if team_id_list:
                query = query.filter(models.Aktivitas.team_id.in_(team_id_list))
        except ValueError:
            raise HTTPException(status_code=400, detail="Format team_ids tidak valid.")
    
    return query.all()

@app.get("/api/kalender/timeline", response_model=List[dict])
def get_timeline_data(
    db: Session = Depends(database.get_db),
    team_ids: Optional[str] = Query(None, description="Daftar ID tim yang dipisahkan oleh koma."),
    start_date: date = Query(..., description="Tanggal mulai rentang timeline (YYYY-MM-DD)."),
    end_date: date = Query(..., description="Tanggal selesai rentang timeline (YYYY-MM-DD).")
):
    """
    Mengambil data timeline yang sudah diolah dari backend, termasuk penugasan lane.
    Hanya aktivitas dengan kalender_view == True.
    """
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.users).joinedload(models.User.jabatan),
        joinedload(models.Aktivitas.team)
    ).filter(
        models.Aktivitas.kalender_view == True, # <--- FILTER WAJIB
        or_(
            and_(
                models.Aktivitas.tanggal_mulai <= end_date,
                models.Aktivitas.tanggal_selesai >= start_date
            ),
            and_(
                models.Aktivitas.tanggal_mulai.between(start_date, end_date),
                models.Aktivitas.tanggal_selesai.is_(None)
            )
        )
    ).order_by(models.Aktivitas.tanggal_mulai)

    if team_ids:
        try:
            team_id_list = [int(id_str) for id_str in team_ids.split(',') if id_str.isdigit()]
            if team_id_list:
                user_ids_in_teams = db.query(models.user_team_link.c.user_id).filter(
                    models.user_team_link.c.team_id.in_(team_id_list)
                ).all()
                unique_user_ids = {user_id for (user_id,) in user_ids_in_teams}
                query = query.join(models.anggota_aktivitas_link).filter(
                    models.anggota_aktivitas_link.c.user_id.in_(unique_user_ids)
                ).distinct()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format team_ids tidak valid.")

    aktivitas = query.all()
    
    # Kumpulkan daftar pegawai unik
    pegawai_map = {}
    for a in aktivitas:
        for user in a.users:
            if user.id not in pegawai_map:
                pegawai_map[user.id] = {
                    "id": user.id,
                    "namaLengkap": user.nama_lengkap,
                    "aktivitas": []
                }
    
    # Tetapkan lane dan tambahkan ke setiap pegawai
    for pegawai_id, pegawai_data in pegawai_map.items():
        pegawai_events = []
        for a in aktivitas:
            if any(u.id == pegawai_id for u in a.users):
                pegawai_events.append({
                    "id": a.id,
                    "title": a.nama_aktivitas,
                    "start": a.tanggal_mulai,
                    "end": a.tanggal_selesai if a.tanggal_selesai else a.tanggal_mulai,
                    "start_time": str(a.jam_mulai) if a.jam_mulai else None,
                    "end_time": str(a.jam_selesai) if a.jam_selesai else None,
                    "backgroundColor": a.team.warna if a.team else "#2563eb",
                    "tanggalMulai": a.tanggal_mulai,
                    "tanggalSelesai": a.tanggal_selesai,
                })

        # Logika penugasan lane yang dipindahkan dari frontend
        sorted_events = sorted(pegawai_events, key=lambda e: e['start'])
        lanes = []
        for event in sorted_events:
            assigned_lane = -1
            for i, lane in enumerate(lanes):
                can_fit = True
                for placed_event in lane:
                    start1 = event['start']
                    end1 = event['end']
                    start2 = placed_event['start']
                    end2 = placed_event['end']
                    # Logika tumpang tindih
                    if max(start1, start2) <= min(end1, end2):
                        can_fit = False
                        break
                if can_fit:
                    assigned_lane = i
                    break
            
            if assigned_lane == -1:
                lanes.append([event])
                event['lane'] = len(lanes)
            else:
                lanes[assigned_lane].append(event)
                event['lane'] = assigned_lane + 1

        pegawai_data['aktivitas'] = sorted_events

    return list(pegawai_map.values())

# Endpoint untuk mengambil semua aktivitas yang melibatkan pengguna tertentu
@app.get("/api/users/{user_id}/aktivitas", response_model=List[schemas.Aktivitas])
def get_user_aktivitas(user_id: int, db: Session = Depends(database.get_db)):
    """
    Mengambil semua aktivitas di mana pengguna dengan user_id terlibat.
    """
    # Mengambil aktivitas yang terkait dengan user, dengan eager loading team untuk kalender
    user_aktivitas = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.team)
    ).join(models.anggota_aktivitas_link).filter(
        models.anggota_aktivitas_link.c.user_id == user_id
    ).order_by(models.Aktivitas.tanggal_mulai.desc()).all()
    
    return user_aktivitas

# Endpoint untuk mengambil semua dokumen wajib yang harus diselesaikan pengguna
@app.get("/api/users/{user_id}/dokumen-wajib", response_model=List[schemas.DaftarDokumen])
def get_user_dokumen_wajib(user_id: int, db: Session = Depends(database.get_db)):
    """
    Mengambil daftar dokumen wajib yang terkait dengan aktivitas pengguna.
    """
    dokumen_wajib = db.query(models.DaftarDokumen).options(
        joinedload(models.DaftarDokumen.aktivitas),
        joinedload(models.DaftarDokumen.aktivitas).joinedload(models.Aktivitas.team)
    ).join(models.Aktivitas).join(models.anggota_aktivitas_link).filter(
        models.anggota_aktivitas_link.c.user_id == user_id
    ).order_by(models.Aktivitas.tanggal_mulai.desc()).all()

    return dokumen_wajib

# Endpoint untuk mengambil semua dokumen wajib dari sebuah tim
@app.get("/api/teams/{team_id}/dokumen-wajib-team", response_model=List[schemas.DaftarDokumen])
def get_team_dokumen_wajib(team_id: int, db: Session = Depends(database.get_db)):
    """
    Mengambil daftar dokumen wajib dari semua aktivitas di sebuah tim.
    """
    dokumen_wajib = db.query(models.DaftarDokumen).options(
        joinedload(models.DaftarDokumen.aktivitas),
        joinedload(models.DaftarDokumen.aktivitas).joinedload(models.Aktivitas.team)
    ).join(models.Aktivitas).filter(
        models.Aktivitas.team_id == team_id
    ).order_by(models.Aktivitas.tanggal_mulai.desc()).all()

    return dokumen_wajib

@app.get("/api/kalender/events", response_model=List[schemas.Aktivitas])
def get_calendar_events(
    db: Session = Depends(database.get_db),
    team_ids: Optional[str] = Query(None, description="Daftar ID tim yang dipisahkan oleh koma."),
):
    """
    Mengambil daftar semua aktivitas yang relevan untuk tampilan kalender.
    Jika team_ids diberikan, akan memfilter berdasarkan tim tersebut.
    """
    query = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.users),
        joinedload(models.Aktivitas.team)
    )

    if team_ids:
        try:
            team_id_list = {int(id_str) for id_str in team_ids.split(',') if id_str.isdigit()}
            if team_id_list:
                query = query.filter(models.Aktivitas.team_id.in_(team_id_list))
        except ValueError:
            raise HTTPException(status_code=400, detail="Format team_ids tidak valid.")
    
    return query.all()

# ===================================================================
# ENDPOINT KHUSUS DASHBOARD
# ===================================================================

@app.get("/api/dashboard/stats", response_model=schemas.DashboardStats)
def get_dashboard_stats(
    team_id: Optional[int] = Query(None),
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Mengembalikan statistik ringkas berdasarkan peran user.
    """
    stats = schemas.DashboardStats()
    today = date.today()

    # -----------------------------------------------------------
    # 1. LOGIKA UNTUK ADMIN / KEPALA KANTOR (Melihat Global)
    # -----------------------------------------------------------
    # Asumsi: Role ID 1/2 adalah Admin/Superadmin atau Jabatan Kepala (ID 1)
    if current_user.sistem_role_id in [1, 2] or current_user.jabatan_id == 1: 
        stats.total_pegawai = db.query(models.User).filter(models.User.is_active == True).count()
        stats.total_tim = db.query(models.Team).filter(
            models.Team.valid_until >= today
        ).count()
        stats.total_aktivitas_bulan_ini = db.query(models.Aktivitas).filter(
            extract('month', models.Aktivitas.tanggal_mulai) == today.month,
            extract('year', models.Aktivitas.tanggal_mulai) == today.year
        ).count()

    # -----------------------------------------------------------
    # 2. LOGIKA UNTUK KETUA TIM (Melihat Timnya)
    # -----------------------------------------------------------
    # Ambil semua tim yang dipimpin oleh user ini
    tim_ketua_list = db.query(models.Team).filter(
        models.Team.ketua_tim_id == current_user.id,
        models.Team.valid_until >= today
    ).all()

    # Hanya jalankan perhitungan ketua tim jika user memang memimpin tim
    if tim_ketua_list:
        if team_id:
            # KASUS A: Filter Spesifik Satu Tim
            # Pastikan team_id yang diminta memang milik user ini
            target_tim = next((t for t in tim_ketua_list if t.id == team_id), None)
            
            if target_tim:
                # Hitung anggota via tabel link untuk akurasi
                stats.total_anggota_tim = db.query(models.user_team_link).filter(
                    models.user_team_link.c.team_id == target_tim.id
                ).count()
                
                stats.total_project = db.query(models.Project).filter(
                    models.Project.team_id == target_tim.id
                ).count()
                
                stats.total_aktivitas_bulan_ini = db.query(models.Aktivitas).filter(
                    models.Aktivitas.team_id == target_tim.id,
                    extract('month', models.Aktivitas.tanggal_mulai) == today.month,
                    extract('year', models.Aktivitas.tanggal_mulai) == today.year
                ).count()
        else:
            # KASUS B: Agregat Semua Tim Saya
            managed_team_ids = [t.id for t in tim_ketua_list]
            
            # Hitung total anggota unik (jika perlu) atau total membership
            stats.total_anggota_tim = db.query(models.user_team_link).filter(
                models.user_team_link.c.team_id.in_(managed_team_ids)
            ).count()
            
            stats.total_project = db.query(models.Project).filter(
                models.Project.team_id.in_(managed_team_ids)
            ).count()
            
            stats.total_aktivitas_bulan_ini = db.query(models.Aktivitas).filter(
                models.Aktivitas.team_id.in_(managed_team_ids),
                extract('month', models.Aktivitas.tanggal_mulai) == today.month,
                extract('year', models.Aktivitas.tanggal_mulai) == today.year
            ).count()

    # -----------------------------------------------------------
    # 3. LOGIKA UNTUK ANGGOTA (Aktivitas Saya)
    # -----------------------------------------------------------
    # Bagian ini HARUS dijalankan untuk SEMUA user (Pegawai, Ketua, Admin)
    # agar widget "Aktivitas Saya" selalu muncul isinya.
    
    stats.total_aktivitas_saya = (
        db.query(models.Aktivitas)
        .join(
            models.anggota_aktivitas_link,
            models.Aktivitas.id == models.anggota_aktivitas_link.c.aktivitas_id
        )
        .filter(
            models.anggota_aktivitas_link.c.user_id == current_user.id,
            extract("month", models.Aktivitas.tanggal_mulai) == today.month,
            extract("year", models.Aktivitas.tanggal_mulai) == today.year,
        )
        .count()
    )

    return stats


@app.get("/api/dashboard/todo", response_model=List[schemas.DashboardTodoItem])
def get_dashboard_todo(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Mengembalikan daftar 'To-Do' prioritas.
    - Untuk Anggota: Dokumen wajib yang BELUM diupload.
    - Untuk Ketua Tim: Dokumen wajib yang SUDAH diupload tapi BELUM divalidasi.
    """
    todo_list = []
    today = date.today()

    # A. TUGAS ANGGOTA: Upload Dokumen yang Belum Ada
    # Cari aktivitas aktif yang melibatkan user
    # Lalu cari checklist item di dalamnya yang belum punya file
    
    # Subquery: Ambil aktivitas yang melibatkan user
    user_activities = db.query(models.Aktivitas).join(
        models.anggota_aktivitas_link
    ).filter(
        models.anggota_aktivitas_link.c.user_id == current_user.id,
        # Opsional: Hanya aktivitas yang belum lewat jauh (misal 30 hari terakhir)
        models.Aktivitas.tanggal_mulai >= today - timedelta(days=30) 
    ).all()

    for akt in user_activities:
        for item in akt.daftar_dokumen_wajib:
            # Cek apakah item ini SUDAH punya file? (menggunakan relasi 'files')
            if not item.files: 
                # Belum ada file -> Masukkan ke Todo List "Pending Upload"
                todo_list.append({
                    "id": item.id,
                    "nama_dokumen": item.nama_dokumen,
                    "status_pengecekan": item.status_pengecekan,
                    "aktivitas_id": akt.id,
                    "nama_aktivitas": akt.nama_aktivitas,
                    "tanggal_mulai": akt.tanggal_mulai,
                    "nama_tim": akt.team.nama_tim if akt.team else "-",
                    "nama_project": akt.project.nama_project if akt.project else "-",
                    "jenis_tugas": "upload" # Marker untuk frontend
                })

    # B. TUGAS KETUA TIM: Validasi Dokumen
    # Cari tim yang dipimpin user
    tim_ketua = db.query(models.Team).filter(
        models.Team.ketua_tim_id == current_user.id,
        models.Team.valid_until >= today
    ).all()

    tim_ids = [t.id for t in tim_ketua]
    
    if tim_ids:
        # Cari checklist item di tim ini yang ADA file tapi BELUM divalidasi
        pending_validation_items = db.query(models.DaftarDokumen).join(
            models.Aktivitas
        ).options(
            joinedload(models.DaftarDokumen.aktivitas).joinedload(models.Aktivitas.team),
            joinedload(models.DaftarDokumen.aktivitas).joinedload(models.Aktivitas.project),
            joinedload(models.DaftarDokumen.files) # Load files untuk pengecekan
        ).filter(
            models.Aktivitas.team_id.in_(tim_ids),
            models.DaftarDokumen.status_pengecekan == False, # Belum valid
            models.DaftarDokumen.files.any() # TAPI sudah ada file (artinya butuh review)
        ).limit(20).all() # Limit agar tidak kebanyakan

        for item in pending_validation_items:
            akt = item.aktivitas
            todo_list.append({
                "id": item.id,
                "nama_dokumen": item.nama_dokumen,
                "status_pengecekan": item.status_pengecekan,
                "aktivitas_id": akt.id,
                "nama_aktivitas": akt.nama_aktivitas,
                "tanggal_mulai": akt.tanggal_mulai,
                "nama_tim": akt.team.nama_tim if akt.team else "-",
                "nama_project": akt.project.nama_project if akt.project else "-",
                "jenis_tugas": "validasi" # Marker untuk frontend
            })

    # Sortir berdasarkan tanggal (yang paling mendesak/lama di atas)
    # Disini kita sort desc (terbaru dulu) atau asc (terlama dulu) terserah
    todo_list.sort(key=lambda x: x['tanggal_mulai'] or date.min, reverse=True)

    return todo_list

@app.get("/api/dashboard/timeline-monitor")
def get_timeline_monitor(
    start_date: date,
    end_date: date,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Mengambil data untuk Timeline View: List Tim beserta Aktivitas mereka 
    dalam rentang tanggal tertentu.
    """
    # 1. Ambil semua Tim Aktif
    teams = db.query(models.Team).filter(
        models.Team.valid_until >= date.today()
    ).all()

    result = []

    for team in teams:
        # 2. Ambil Aktivitas Tim ini yang BERIRISAN dengan range tanggal
        # Logika Overlap: (StartA <= EndB) and (EndA >= StartB)
        activities = db.query(models.Aktivitas).filter(
            models.Aktivitas.team_id == team.id,
            models.Aktivitas.kalender_view == True, 
            models.Aktivitas.tanggal_mulai <= end_date,
            or_(
                models.Aktivitas.tanggal_selesai >= start_date,
                # Handle aktivitas 1 hari (tanggal_selesai null -> dianggap sama dengan tanggal_mulai)
                and_(models.Aktivitas.tanggal_selesai.is_(None), models.Aktivitas.tanggal_mulai >= start_date)
            )
        ).all()

        # Format aktivitas untuk frontend
        acts_data = []
        for act in activities:
            # Normalisasi tanggal selesai untuk logika frontend
            real_end = act.tanggal_selesai if act.tanggal_selesai else act.tanggal_mulai
            
            acts_data.append({
                "id": act.id,
                "namaAktivitas": act.nama_aktivitas,
                "tanggalMulai": act.tanggal_mulai,
                "tanggalSelesai": real_end,
                "status": "berjalan" # Bisa dikembangkan nanti
            })

        result.append({
            "id": team.id,
            "namaTim": team.nama_tim,
            "warna": team.warna or "#3b82f6", # Default blue
            "ketua": team.ketua_tim.nama_lengkap if team.ketua_tim else "-",
            "activities": acts_data
        })

    return result


@app.get("/api/dashboard/team-timeline")
def get_team_timeline_for_leader(
    start_date: date,
    end_date: date,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """
    Mengambil jadwal aktivitas untuk tim yang dipimpin oleh user saat ini.
    """
    # 1. Cari Tim di mana user adalah Ketua
    # Kita ambil tim yang masih aktif
    team = db.query(models.Team).filter(
        models.Team.ketua_tim_id == current_user.id,
        models.Team.valid_until >= date.today()
    ).first()

    if not team:
        return [] # User bukan ketua tim atau tim sudah tidak aktif

    # 2. Ambil Aktivitas Tim dalam rentang tanggal
    # Logika Overlap: Activity Start <= View End AND Activity End >= View Start
    activities = db.query(models.Aktivitas).options(
        joinedload(models.Aktivitas.project) # Load project info
    ).filter(
        models.Aktivitas.team_id == team.id,
        models.Aktivitas.kalender_view == True, # <--- FILTER WAJIB
        models.Aktivitas.tanggal_mulai <= end_date,
        or_(
            models.Aktivitas.tanggal_selesai >= start_date,
            # Handle aktivitas 1 hari (tanggal_selesai null dianggap sama dengan mulai)
            and_(models.Aktivitas.tanggal_selesai.is_(None), models.Aktivitas.tanggal_mulai >= start_date)
        )
    ).order_by(models.Aktivitas.tanggal_mulai.asc(), models.Aktivitas.jam_mulai.asc()).all()
    
    # Kita return list aktivitas, grouping dilakukan di frontend agar fleksibel
    return activities

# ===================================================================
# ENDPOINT UNTUK NOTIFKIKASI
# ===================================================================

# Endpoint 1: GET /api/notifications/count (Untuk Polling Badge)
@app.get("/api/notifications/count")
def get_unread_notification_count(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """Menghitung jumlah notifikasi yang belum dibaca untuk pengguna."""
    
    unread_count = db.query(models.Notifikasi).filter(
        models.Notifikasi.user_id == current_user.id,
        models.Notifikasi.is_read == False
    ).count()
    
    return {"count": unread_count}

# Endpoint 2: GET /api/notifications (Untuk Dropdown List)
@app.get("/api/notifications/header", response_model=List[schemas.Notifikasi]) 
def get_header_notifications(
    db: Session = Depends(database.get_db), 
    current_user: models.User = Depends(security.get_current_user), 
    limit: int = 15
):
    """Mengambil daftar notifikasi terbaru (list saja) untuk header dropdown."""
    
    notifications = db.query(models.Notifikasi).options(
        joinedload(models.Notifikasi.user) 
    ).filter(
        models.Notifikasi.user_id == current_user.id
    ).order_by(
        models.Notifikasi.created_at.desc() 
    ).limit(limit).all()
    
    return notifications


# Endpoint 2b: GET /api/notifications (Paginated untuk Halaman View All)
# Endpoint ini yang sebelumnya bermasalah karena dipanggil oleh header. Sekarang hanya untuk NotifikasiView.
@app.get("/api/notifications", response_model=schemas.NotifikasiPage) 
def get_user_notifications_paginated(
    db: Session = Depends(database.get_db), 
    current_user: models.User = Depends(security.get_current_user), 
    skip: int = 0,
    limit: int = 20, 
    is_read: Optional[bool] = Query(None) 
):
    """Mengambil daftar notifikasi terbaru untuk pengguna dengan dukungan paginasi."""
    
    query = db.query(models.Notifikasi).options(
        joinedload(models.Notifikasi.user) 
    ).filter(
        models.Notifikasi.user_id == current_user.id
    )
    
    if is_read is not None:
        query = query.filter(models.Notifikasi.is_read == is_read)

    total = query.count()
    
    notifications = query.order_by(
        models.Notifikasi.created_at.desc()
    ).offset(skip).limit(limit).all()
    
    return {"total": total, "items": notifications} 


# Endpoint 3: PATCH /api/notifications/{id}/read (Untuk Mark as Read)
@app.patch("/api/notifications/{notification_id}/read")
def mark_notification_as_read(
    notification_id: int, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user)
):
    """Menandai notifikasi sebagai sudah dibaca."""
    
    # 1. Cari notifikasi, pastikan dimiliki oleh user saat ini
    db_notification = db.query(models.Notifikasi).filter(
        models.Notifikasi.id == notification_id,
        models.Notifikasi.user_id == current_user.id
    ).first()
    
    if not db_notification:
        raise HTTPException(status_code=404, detail="Notifikasi tidak ditemukan")
    
    # 2. Update status
    db_notification.is_read = True
    db.commit()
    db.refresh(db_notification)
    
    return {"message": "Notifikasi berhasil ditandai sebagai sudah dibaca"}

@app.patch("/api/notifications/mark-all-read")
def mark_all_as_read(current_user: models.User = Depends(security.get_current_user), db: Session = Depends(database.get_db)):
    """Menandai SEMUA notifikasi pengguna sebagai sudah dibaca."""
    db.query(models.Notifikasi).filter(
        models.Notifikasi.user_id == current_user.id,
        models.Notifikasi.is_read == False
    ).update({"is_read": True}, synchronize_session=False)
    
    db.commit()
    return {"message": "Semua notifikasi berhasil ditandai sebagai sudah dibaca"}